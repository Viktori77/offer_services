from aiogram import Router, F, Bot
from aiogram.filters import CommandStart,StateFilter
from aiogram.types import Message, CallbackQuery, FSInputFile, ReplyKeyboardRemove, ContentType, BufferedInputFile
from aiogram.exceptions import TelegramForbiddenError
from datetime import time, datetime
from aiogram.fsm.state import StatesGroup, State
from sqlalchemy import func
from aiogram.fsm.context import FSMContext
from database.db_handlers import (
    add_or_get_user, get_prefix_models, get_models, update_event_action, get_event_by_id, get_user_by_tg_id, update_user_ban_status, get_events_by_date_and_town, get_events_future_by_town, update_grade, get_all_records, get_all_in_town, add_or_update_record, remove_record, get_model_columns, get_record_counts_for_models, get_model_by_name, get_record_by_id, get_record_count_analog, update_message_status, get_records_for_user, get_tg_ids_from_model, add_or_update_file_record, get_template, get_records_for_user_in_settings, add_record, remove_records_all, clear_booking_fields, copy_record_with_empty_user_data, get_numberPhone, update_numberphone, update_book_agree_status
)
from helpers.messages import create_new_message_text, send_messages_to_users_all, format_data_message, generate_message_from_model, create_schedule_message, create_user_message, get_full_info_message, format_message_for_send, send_message_to_user, get_full_info_message_data, format_data_one_message, get_message_bookig
from bot.create_bot import admins, url_recording
import bot.keyboards.all_keyboards as kb
from openpyxl import load_workbook
from io import BytesIO
import os
from struction import combined_model_names_for_admin, combined_model_names_for_users, sectionBus, all_sections, combined_model_names_for_settings, reversed_all_town_name, analog_model_names, analog_model_names_reserv
from utils import get_current_time, remove_settings_prefix, get_today_date_dmy, get_tomorrow_date, get_filters_for_model, prepare_model_data, get_section_name_ru, convert_orm_to_dict, get_places_from_events
from helpers.valid import is_valid_date, is_valid_time, is_valid_price
from helpers.create_file import create_excel_file
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

router = Router()


# Состояния для добавления города
class AddTown(StatesGroup):
    waiting_for_town_name = State()

# Состояния для добавления события
class AddEvent(StatesGroup):
    waiting_for_event_town = State()
    waiting_for_event_name = State()
    waiting_for_event_description = State()
    waiting_for_event_date = State()
    waiting_for_event_time = State()
    waiting_for_link_to_source = State()
    waiting_for_price = State()
    waiting_for_place = State()

# Состояния для добавления расписания
class AddBusSchedule(StatesGroup):
    waiting_for_town = State()  # Город
    waiting_for_section_bus = State()  # Раздел
    waiting_for_number_bus = State()  # Раздел
    waiting_for_start_place = State()  # Откуда идет
    waiting_for_finish_place = State()  # Куда идет
    waiting_for_time_start = State()  # Время старта
    waiting_for_time_finish = State()  # Время финиша
    waiting_for_days = State()  # Дни, когда ездит
    waiting_for_link_to_source = State()  # Ссылка на источник


# Состояния для удаления города
class RemoveTown(StatesGroup):
    waiting_for_town_name = State()   

# Состояния для удаления события по ID
# class RemoveEvent(StatesGroup):
#     waiting_for_event_id = State()

class RemoveModel(StatesGroup):
    waiting_for_id = State()  # Состояние для ожидания ввода ID
    model_name = None  # Поле для хранения названия модели


# Состояния для обновления статуса события
class UpdateEventAction(StatesGroup):
    waiting_for_event_id = State()
    waiting_for_new_action = State()

class BanUser(StatesGroup):
    waiting_for_user_id = State()  # Ожидание ввода ID пользователя
    waiting_for_ban_action = State()  # Ожидание выбора действия (бан/разбан)

# Определяем состояние для ввода периода
class EventPeriod(StatesGroup):
    waiting_for_start_date = State()
    waiting_for_end_date = State()

class UploadData(StatesGroup):
    waiting_for_model = State()  # Ожидание выбора модели
    waiting_for_file = State()   # Ожидание загрузки файла

class UpdateGrade(StatesGroup):
    waiting_for_model = State()  # Ожидание выбора модели
    waiting_for_id = State()     # Ожидание ввода ID строки
    waiting_for_grade = State()  # Ожидание ввода нового грейда

class AddAnalogModel(StatesGroup):
    waiting_for_town = State()  # выбор города
    waiting_for_model_name = State()  # Ожидание выбора модели
    waiting_for_option = State()   # Ожидание выбора подраздела (опции)
    waiting_for_name = State()  # Ожидание ввода названия
    waiting_for_description_small = State()  # Ожидание ввода краткого описания
    waiting_for_description_full = State()  # Ожидание ввода полного описания
    waiting_for_schedule = State()  # Ожидание ввода графика работы
    waiting_for_coordinates = State()  # Ожидание ввода координат
    waiting_for_address = State()  # Ожидание ввода адреса
    waiting_for_phone = State()  # Ожидание ввода телефона
    waiting_for_website = State()  # Ожидание ввода сайта
    waiting_for_nameUser = State()  # Ожидание ввода информации о том, кто предложил
    waiting_for_tgId = State()  # Ожидание ввода
    waiting_for_confirmation = State()  # Ожидание ввода грейда
    waiting_for_reject_reason = State()  # Состояние ожидания текста объяснения

class SendMessageState(StatesGroup):
    CHOOSE_RECIPIENT = State()  # Выбор получателя (всем или одному)
    ENTER_TOPIC = State()  # Ввод темы сообщения
    ENTER_MESSAGE = State()     # Ввод текста сообщения
    ENTER_USER_ID = State()     # Ввод ID пользователя (если выбран "одному")

class ReplyToMessage(StatesGroup):
    waiting_for_reply_text = State()  # Состояние ожидания текста ответа

class EventState(StatesGroup):
    event_date = State()
    town = State()
    # booking_date = State()
    # booking_time = State()
    # confirmation = State()
    # success = State()

# Создаем состояние для ожидания файла
class Form(StatesGroup):
    waiting_for_template = State()
    waiting_for_ready = State()

    
class TemplateRecord(StatesGroup):
    waiting_for_model = State()  # Ожидание выбора модели
    waiting_for_file = State()   # Ожидание загрузки файла


# Хэндлер для команды /start
@router.message(CommandStart())
async def cmd_start(message: Message, bot: Bot, state: FSMContext):
    """
    Обрабатывает команду /start и показывает основную клавиатуру.
    """
    # Очищаем текущее состояние пользователя
    await state.clear()

    # Получаем данные пользователя
    tg_id = int(message.from_user.id)
    tg_name = message.from_user.username or message.from_user.first_name
    time_reg_obj=get_current_time()
    time_reg=time_reg_obj.strftime("%Y-%m-%d %H:%M:%S")

    # Получаем количество пользователей
    users=await get_all_records("User")
    count_users=len(users)

    # Получаем количество предложений
    # offers=await get_record_counts_for_models(analog_model_names)
    # count_offers=(sum(offers.values())-1-1-4-22-2-16-6-1-1-1)

    # Проверяем есть ли пользовать в базе
    check_user = await get_user_by_tg_id(tg_id)
    if check_user:
        logger.info(f'Пользователь {tg_id} уже есть в базе')
    else:
        # Отправляем сообщение админу c информацией о новом пользователе
        for admin_id in admins:
            try:
                await bot.send_message(
                    admin_id,
                    f"Новый пользователь: {tg_name}"
                )
            except TelegramForbiddenError as tg_error:
                logger.error(f"Не удалось отправить сообщение админу {admin_id}: {tg_error}")


    # # Добавляем пользователя в базу данных
    try:
        user = await add_or_get_user(tg_id, tg_name, time_reg)
    except Exception as e:
        logger.error(f"Ошибка при добавления пользователя с тг id: {tg_id}: {e}")
        return

    if tg_id in admins:
        keyboard=kb.main_admin
    else:      
        keyboard=kb.main_users
            
    await message.answer(
                        f"Добро пожаловать, {message.from_user.first_name}!\n\n"
                                
                        "✅ Где покушать?\n"
                        "✅ Куда сходить вечером?\n"
                        "✅ Где найти шар для гендер-пати?\n"
                        "✅ Где отметить день Рождения?\n"
                        "✅ Где купить?\n"
                        "✅ На какую секцию записать своего ребенка?\n\n"

                        "🎉 Eсли Вы организуете мастер-классы, например, или массовые спортивные мероприятия, предположим, то можно добавить свою афишу на месяц.\n"
                        "🚌 Если Вы частная организация по перевозке людей в городе, то можете добавить расписание рейсов.\n"
                        f"🏠 Eсли Вы владелец какой-либо организации или просто оказываете услуги, о которых должны знать люди, то смело добавляйтесь в соответствующий раздел. 🏆 И тогда у Вас откроется возможность пользоваться платформой Запишись. {url_recording}.\n\n"
                        f"✔️ У Вас появится свой личный кабинет, в котором Вы сможете добавлять расписания, а люди будут иметь доступ к нему по ссылке: {url_recording} в любое время дня и ночи. Теперь эта рутина с записями останется за нами, а Вам всего лишь будет приходить оповещение, что у Вас новая запись. А если Ваш клиент еще и является пользователем данного бота, то он будет получать напоминание о сделанной записи с просьбой подтвердить. Ответ от клиента будет приходить Вам. То есть теперь Вам можно не заморачиваться о напоминаниях, за Вас это сделает супер-бот. 🎁💻\n\n"

                        '🧱 Вместе построим лучший город, в котором люди всегда будут знать "что? где? когда?"\n\n'

                        f'👤 Количество пользователей: {count_users}.\n'
                        # f'🏆 Количество предложений: {count_offers}.\n\n'
                        ,
                        reply_markup=keyboard,
                        disable_web_page_preview=True,
                        parse_mode="HTML"
                    )
        

      
@router.callback_query(F.data == "my_settings")
async def handle_my_settings(callback: CallbackQuery):
    """
    Обрабатывает нажатие на кнопку "Настройки уведомлений".
    """
    # Получаем список моделей, начинающихся на 'settings'
    models = get_prefix_models()

    # Создаем клавиатуру
    keyboard = kb.create_settings_keyboard(models, combined_model_names_for_settings)

    # Отправляем сообщение с клавиатурой
    await callback.message.answer("Выберите раздел или город, в котором хотите отключить или включить уведомления:", reply_markup=keyboard)

    # Подтверждаем обработку callback
    await callback.answer()

@router.callback_query(F.data == "disable_all_notifications")
async def handle_settings_choice(callback: CallbackQuery, state: FSMContext):
    """
    Обрабатывает выбор настройки вкл выкл все уведомления".
    """
    if callback.data == "disable_all_notifications":
        # Обработка кнопки "Отключить все уведомления"
        tg_id = callback.from_user.id
        # models = get_all_settings_models()  # Функция, которая возвращает все модели, начинающиеся на 'settings_'

        for model in combined_model_names_for_users:
            model=f'settings_{model}'
            check=await get_records_for_user_in_settings(tg_id=tg_id, model_name=model)
            filters={"tg_id": tg_id}
            if check:  # Проверяем, есть ли пользователь в таблице
                result = await remove_record(filters=filters, model_name=model)  # Удаляем пользователя из таблицы
            else:
                data = {
                    "tg_id": tg_id,            
                    }
                result = await add_or_update_record(model, filters, data)
                # await add_user_to_model(tg_id, model)  # Добавляем пользователя в таблицу
        if result:
            # Отправляем сообщение об успешном выполнении
            await callback.message.answer("Все уведомления были включены." if check else "Все уведомления были отключены.")
            return
        else:
            logger.error(f'Все уведомления не могут быть отключены\включены у пользователя: {tg_id}')

@router.callback_query(F.data.startswith("settings_"))
async def handle_settings_choice(callback: CallbackQuery, state: FSMContext):
    """
    Обрабатывает выбор настройки и показывает клавиатуру с кнопками "ВКЛ" и "ВЫКЛ".
    """
    # if callback.data == "disable_all_notifications":
    #     # Обработка кнопки "Отключить все уведомления"
    #     tg_id = callback.from_user.id
    #     # models = get_all_settings_models()  # Функция, которая возвращает все модели, начинающиеся на 'settings_'

    #     for model in combined_model_names_for_users:
    #         model=f'settings_{model}'
    #         check=await get_records_for_user_in_settings(tg_id=tg_id, model_name=model)
    #         filters={"tg_id": tg_id}
    #         if check:  # Проверяем, есть ли пользователь в таблице
    #             result = await remove_record(filters=filters, model_name=model)  # Удаляем пользователя из таблицы
    #         else:
    #             data = {
    #                 "tg_id": tg_id,            
    #                 }
    #             result = await add_or_update_record(model, filters, data)
    #             # await add_user_to_model(tg_id, model)  # Добавляем пользователя в таблицу
    #     if result:
    #         # Отправляем сообщение об успешном выполнении
    #         await callback.message.answer("Все уведомления были включены." if check else "Все уведомления были отключены.")
    #         return
    #     else:
    #         logger.error(f'Все уведомления не могут быть отключены\включены у пользователя: {tg_id}')
    
    # Извлекаем название модели из callback_data
    model_name = remove_settings_prefix(callback.data)
    logger.info(f'model_name: {model_name}')
    if len(model_name.split('_'))==1:
        # Получаем перевод названия модели из словаря
        model_name_translate = combined_model_names_for_settings.get(model_name, model_name)  # Если перевод не найден, используем оригинальное имя
        logger.info(f'model_name_translate: {model_name_translate}')

        # Создаем клавиатуру с кнопками "ВКЛ" и "ВЫКЛ"
        keyboard = kb.create_on_off_keyboard(model_name)

        # Отправляем сообщение с клавиатурой
        await callback.message.answer(
            f'Включите или выключите уведомления для раздела "{model_name_translate}":',
            reply_markup=keyboard
        )
    else:
        # Извлекаем название модели из callback_data
        data = remove_settings_prefix(callback.data).split("_")
        
        model_name=data[0]
        
        town=data[2]
        

        # # Получаем перевод названия модели для настройки города из словаря
        model_name_eu = reversed_all_town_name.get(town, town)  # Если перевод не найден, используем оригинальное имя
        logger.info(f'model_name_eu: {model_name_eu}')

        # Инициализируем список исключенных tg_id
        excluded_tg_ids = set()

        # получаем список tg_id из указанной модели
        if model_name_eu is not None:
            settings_model_name = f"settings_{model_name_eu}"
            try:
                excluded_tg_ids = await get_tg_ids_from_model(settings_model_name)
            except Exception as e:
                logger.error(f"Ошибка при получении данных из модели {settings_model_name}: {e}")
                return
        else:
            logger.error(f'Модель: {model_name_eu} не найдена')
        

        #Проверяем есть ли пользователь в таблице Город_сеттинг
        tgId=callback.from_user.id
        
        if tgId in excluded_tg_ids:
            # Отправляем сообщение
            await callback.message.answer(
                f'У вас выключены уведомления для города {town}. Включите их, чтобы настроить уведомления по месту проведения'
            )
            return
        else:
            event_id=int(data[1])
            # Получаем мероприятие по id
            event_first = await get_event_by_id(event_id)
            # Получаем место мероприятие по id
            place=event_first.place
            # Сохраняем place в состояние
            await state.update_data(place=place)

            # Создаем клавиатуру с кнопками "ВКЛ" и "ВЫКЛ"
            keyboard = kb.create_on_off_keyboard(model_name=model_name_eu, event_id=event_id)

            # Отправляем сообщение с клавиатурой
            await callback.message.answer(
                f'Включите или выключите уведомления в городе {town} для места проведения: "{place}":',
                reply_markup=keyboard
            )
            
            # # Отправляем сообщение с клавиатурой
            # await callback.message.answer(
            #     f'Включите или выключите уведомления для раздела "{model_name_eu}":',
            #     reply_markup=keyboard
            # )

    # Подтверждаем обработку callback
    await callback.answer()

@router.callback_query(F.data.startswith("selectLocation_"))
async def handle_settings_choice(callback: CallbackQuery):
    """
    Обрабатывает выбор Выбрать Место в настройках для мероприятий.
    """
    # Извлекаем данные из callback_data
    data = callback.data.split("_")
    model_name = data[1]  # Название модели
    # Получаем перевод названия модели из словаря
    model_name_translate = combined_model_names_for_settings.get(model_name, model_name)  # Если перевод не найден, используем оригинальное имя

    town = model_name_translate
    logger.info(f'model_name: {model_name}')

    logger.info(f'model_name_translate: {model_name_translate}')

    town = model_name_translate
    logger.info(f'town: {town}')

    # Получаем мероприятия
    events = await get_all_in_town(model_name="Event", town=town)

    events_list=[
                {
                    "id": event.id,
                    "town": event.town,
                    "event": event.event,
                    "description": event.description,
                    "event_date": event.event_date,
                    "time": event.time,
                    "link_to_source": event.link_to_source,
                    "price": event.price,
                    "place": event.place,
                    "action": event.action,
                    "grade": event.grade,
                }
                for event in events
            ]

    if events_list:
        # Группируем мероприятия по месту проведения
        places = get_places_from_events(events_list)

        # Создаем клавиатуру с местами проведения
        keyboard = kb.create_events_one_place_keyboard(places=places, town=town, action="settings")

        await callback.message.answer("Выберите место проведения:", reply_markup=keyboard.as_markup())
    else:
        await callback.message.answer("Мероприятий не найдено.")


    # # Отправляем сообщение с клавиатурой
    # await callback.message.answer(
    #     f'Включите или выключите уведомления для раздела "{model_name_translate}":',
    #     reply_markup=keyboard
    # )

    # # Подтверждаем обработку callback
    # await callback.answer()

@router.callback_query(F.data.startswith("toggle_"))
async def handle_toggle_choice(callback: CallbackQuery, state: FSMContext):
    """
    Обрабатывает выбор "ВКЛ" или "ВЫКЛ" и добавляет tg_id в таблицу, если выбрано "ВЫКЛ".
    """
    # Извлекаем данные из callback_data
    data = callback.data.split("_")
    model_name = data[1]  # Название модели
    action = data[2]  # "on" или "off"
    # Получаем tg_id пользователя
    tg_id = int(callback.from_user.id)

    if len(data)==3:
        # Получаем перевод названия модели из словаря
        model_name_translate = combined_model_names_for_settings.get(model_name, model_name)  # Если перевод не найден, используем оригинальное имя

        data = {
            "tg_id": tg_id,            
            }
        filters = {"tg_id": tg_id}  # Условия для поиска существующей записи

        model=f'settings_{model_name}'

        text_off = f'Уведомления для раздела "{model_name_translate}" отключены.'
        text_off_dubble = f'Уведомления для раздела "{model_name_translate}" уже были отключены.'

        text_on = f'Уведомления для раздела "{model_name_translate}" включены.'
        text_on_dubble=f'Уведомления для раздела "{model_name_translate}" уже были включены.'

    else:
        # event_id=data[3]
        # Получаем перевод названия модели из словаря
        town = combined_model_names_for_settings.get(model_name, model_name)  # Если перевод не найден, используем оригинальное имя

        # Получаем место проведения из состояния
        data = await state.get_data()
        place = data.get('place')

        data = {
            "town": town,
            "place_name": place,
            "tgId": tg_id,            
            }
        filters = data  # Условия для поиска существующей записи по всем полям

        model='Place_settings'

        text_off = f'Уведомления в городе {town} для места проведения "{place}" отключены.'
        text_off_dubble = f'Уведомления в городе {town} для места проведения "{place}" уже были отключены.'

        text_on = f'Уведомления в городе {town} для места проведения "{place}" включены.'
        text_on_dubble=f'Уведомления в городе {town} для места проведения "{place}" уже были включены.'
        
    if action == "off":
        try:
            # Добавляем пользователя в таблицу, если выбрано "ВЫКЛ"
            result = await add_or_update_record(model, filters, data)

            if result:
                await callback.message.answer(text_off)
            else:
                await callback.message.answer(text_off_dubble)

        except ValueError as e:
            logger.error(f"Ошибка при добавлении пользователя {tg_id} в таблицу {model_name}: {e}")
            await callback.message.answer("Неизвестная настройка.")
    else:
        # Если выбрано "ВКЛ", удаляем пользователя из таблицы (если он там есть)
        try:
            result = await remove_record(
            model_name=model,  # Передаем модель 
            filters=filters,  # Фильтр по tg_id
            )
            if result:
                await callback.message.answer(text_on)
            else:
                await callback.message.answer(text_on_dubble)

        except ValueError as e:
            logger.error(f"Ошибка при удалении пользователя {tg_id} из таблицы {model_name}: {e}")
            await callback.message.answer("Неизвестная настройка.")


    # Подтверждаем обработку callback
    await callback.answer()


@router.callback_query(F.data == "choose_сity")
async def handle_view_city(callback: CallbackQuery):
    """
    Обрабатывает нажатие на кнопку "Зайти в город".
    """
    try:
        # Получаем список городов из базы данных
        towns = await get_all_records("Town")
        
        # Сортируем города по грейду (по убыванию)
        sorted_towns = sorted(towns, key=lambda town: town.grade)

        # Извлекаем названия городов (предполагается, что у Town есть атрибут `name`)
        town_names = [town.town for town in sorted_towns]
        
        # Создаем клавиатуру с отсортированными городами
        keyboard = kb.create_towns_keyboard(town_names, "view")
        
        # Отправляем сообщение с клавиатурой
        await callback.message.answer("Выберите город:", reply_markup=keyboard)
        
    except Exception as e:
        logger.error(f"Ошибка при обработке callback 'choose_сity': {e}", exc_info=True)
        await callback.answer("Произошла ошибка. Пожалуйста, попробуйте позже.", show_alert=True)

    # Подтверждаем обработку callback
    await callback.answer()


# @router.callback_query(F.data == "add")
@router.callback_query(F.data.startswith("section_for_"))
async def handle_add(callback: CallbackQuery, state: FSMContext):
    """
    Обрабатывает нажатие на кнопку "Добавить данные", "Удалить данные", "Загрузить данные". для админа права только 
    """
    if callback.from_user.id in admins:
        # Получаем список моделей
        models = await get_models(combined_model_names_for_admin)
        action = callback.data.split("_")[2]

        if action=="upload":
            await state.set_state(UploadData.waiting_for_model)

        # Создаем клавиатуру с моделями
        keyboard = kb.create_models_keyboard(models, combined_model_names_for_admin, action)

        # Отправляем сообщение с клавиатурой
        await callback.message.answer("Выберите модель:", reply_markup=keyboard)

        # Подтверждаем обработку callback
        await callback.answer()
    else:
        await callback.message.answer("Извините, у Вас нет прав доступа")

@router.callback_query(F.data.startswith("add_"))
async def handle_add_model(callback: CallbackQuery, state: FSMContext):
    """
    Обрабатывает выбор модели для добавления.
    """
    if callback.from_user.id in admins:

        model_name = callback.data.replace("add_", "")
        await state.update_data(model_name=model_name)

        if model_name == "Town":
            await callback.message.answer("Введите название города:")
            await state.set_state(AddTown.waiting_for_town_name)
        elif model_name == "Event":
            # Получаем список городов
            towns = await get_all_records("Town")

            # Извлекаем названия городов
            town_names = [town.town for town in towns]  # Предполагается, что у модели Town есть атрибут name

            # Создаем клавиатуру с городами
            towns_keyboard = kb.create_towns_text_keyboard(town_names)
            
            # Отправляем сообщение с клавиатурой
            await callback.message.answer(
                "В каком городе пройдет мероприятие, выбери вариант из списка ниже (если не видно, то скрой свою клавиатуру):",
                reply_markup=towns_keyboard
            )
            await state.set_state(AddEvent.waiting_for_event_town)
        else:
            # Получаем список городов
            towns = await get_all_records("Town")

            # Извлекаем названия городов
            town_names = [town.town for town in towns] 

            # Создаем клавиатуру с городами
            towns_keyboard = kb.create_towns_text_keyboard(town_names)
            
            # Отправляем сообщение с клавиатурой
            await callback.message.answer(
                "Выберите город (если не видно, то скрой свою клавиатуру):",
                reply_markup=towns_keyboard
            )
            if model_name == "BusSchedule":
                await state.set_state(AddBusSchedule.waiting_for_town)
            else:
                await state.set_state(AddAnalogModel.waiting_for_town)
            

        await callback.answer()
    else:
        await callback.message.answer("Извините, у Вас нет прав доступа")

        

@router.callback_query(F.data.startswith("remove_"))
async def handle_remove_model(callback: CallbackQuery, state: FSMContext):
    """
    Обрабатывает выбор модели для удаления.
    """
    if callback.from_user.id in admins:
        # Извлекаем название модели из callback_data
        model_name = callback.data.replace("remove_", "")
        await state.update_data(model_name=model_name)  # Сохраняем название модели

        if model_name == "Town":
            # Для модели Town запрашиваем название города
            await callback.message.answer("Введите название города для удаления:")
            await state.set_state(RemoveTown.waiting_for_town_name)
        else:
            # Для всех остальных моделей запрашиваем ID
            await callback.message.answer("Введите ID для удаления:")
            await state.set_state(RemoveModel.waiting_for_id)

        await callback.answer()
    else:
        await callback.message.answer("Извините, у Вас нет прав доступа")

@router.message(AddTown.waiting_for_town_name)
async def handle_add_town_name(message: Message, state: FSMContext):
    """
    Обрабатывает ввод названия города для добавления нового города в базу.
    """
    town = message.text

    try:
        # Пытаемся добавить город
        # result = await add_town(town)
        data = {
            "town": town,            
        }
        filters = {"town": town}  # Условия для поиска существующей записи
        result=await add_or_update_record("Town", filters, data)
        # result = await add_town(town)
        
        if result:
            await message.answer(f"Город '{town}' успешно добавлен!")
        else:
            await message.answer(f"Не удалось добавить город '{town}'.")
    
    except Exception as e:
        # Обрабатываем возможные ошибки
        await message.answer(f"Произошла ошибка при добавлении города: {e}")
    
    finally:
        # Очищаем состояние
        await state.clear()

@router.message(RemoveTown.waiting_for_town_name)
async def handle_remove_town_name(message: Message, state: FSMContext):
    """
    Обрабатывает ввод названия города для удаления города из базы.
    """
    town_name = message.text

    try:
        # Пытаемся добавить город
        result = await remove_record(
            model_name="Town",  # Передаем модель Town
            filters={"town": town_name},  # Фильтр по названию города
            
        )
        
        if result:
            await message.answer(f"Город '{town_name}' успешно удален!")
        else:
            await message.answer(f"Не удалось удалить город '{town_name}'.")
    
    except Exception as e:
        # Обрабатываем возможные ошибки
        await message.answer(f"Произошла ошибка при добавлении города: {e}")
    
    finally:
        # Очищаем состояние
        await state.clear()

@router.message(AddEvent.waiting_for_event_town)
async def handle_add_event_name(message: Message, state: FSMContext):
    """
    Обрабатывает ввод названия города.
    """
    await state.update_data(town=message.text)
    await message.answer("Введите название мероприятия:")
    await state.set_state(AddEvent.waiting_for_event_name)


@router.message(AddEvent.waiting_for_event_name)
async def handle_add_event_name(message: Message, state: FSMContext):
    """
    Обрабатывает ввод названия мероприятия.
    """
    await state.update_data(event=message.text)
    await message.answer("Введите описание мероприятия:")
    await state.set_state(AddEvent.waiting_for_event_description)

@router.message(AddEvent.waiting_for_event_description)
async def handle_add_event_name(message: Message, state: FSMContext):
    """
    Обрабатывает ввод названия мероприятия.
    """
    await state.update_data(description=message.text)
    await message.answer("Введите дату мероприятия (в формате YYYY-MM-DD):")
    await state.set_state(AddEvent.waiting_for_event_date)

@router.message(AddEvent.waiting_for_event_date)
async def handle_add_event_date(message: Message, state: FSMContext):
    """
    Обрабатывает ввод даты мероприятия.
    """
    if not is_valid_date(message.text):
        await message.answer("Некорректный формат даты. Введите дату в формате YYYY-MM-DD:")
        return
    

    event_date = datetime.strptime(message.text, "%Y-%m-%d").date()

    await state.update_data(event_date=event_date)
    await message.answer("Введите время мероприятия (в формате ЧЧ:ММ):")
    await state.set_state(AddEvent.waiting_for_event_time)

@router.message(AddEvent.waiting_for_event_time)
async def handle_add_event_time(message: Message, state: FSMContext):
    """
    Обрабатывает ввод времени мероприятия.
    """
    if not is_valid_time(message.text):
        await message.answer("Некорректный формат времени. Введите время в формате ЧЧ:ММ:")
        return

    await state.update_data(time=message.text)
    await message.answer("Введите ссылку на источник в формате http:// или https://:")
    await state.set_state(AddEvent.waiting_for_link_to_source)

@router.message(AddEvent.waiting_for_link_to_source)
async def handle_add_link_to_source(message: Message, state: FSMContext):
    """
    Обрабатывает ввод ссылки на источник.
    """
    if not message.text.startswith("http://") and not message.text.startswith("https://"):
        await message.answer("Некорректная ссылка. Введите ссылку, начинающуюся с http:// или https://:")
        return

    await state.update_data(link_to_source=message.text)
    await message.answer("Введите цену (число или 'бесплатно'):")
    await state.set_state(AddEvent.waiting_for_price)

@router.message(AddEvent.waiting_for_price)
async def handle_add_price(message: Message, state: FSMContext):
    """
    Обрабатывает ввод цены.
    """
    if not is_valid_price(message.text):
        await message.answer("Некорректный формат цены. Введите число или 'бесплатно':")
        return

    await state.update_data(price=message.text)
    await message.answer("Введите место проведения:")
    await state.set_state(AddEvent.waiting_for_place)

@router.message(AddEvent.waiting_for_place)
async def handle_add_place(message: Message, state: FSMContext, bot: Bot):
    """
    Обрабатывает ввод места проведения.
    """
    await state.update_data(place=message.text)
    
    data = await state.get_data()

    town = data["town"]
    event = data["event"]
    description = data["description"]
    event_date = data["event_date"]
    logger.info(f"Data before query: {data}")
    if isinstance(event_date, str):
        event_date = datetime.strptime(event_date, "%Y-%m-%d").date() 

    logger.info(f"Data before query: {data}")  
    time = data["time"]
    link_to_source = data["link_to_source"]
    price = data["price"]
    place = data["place"]
    

    try:
        # Формируем данные для добавления/обновления
        # Формируем данные для добавления/обновления
        event_data = prepare_model_data(
            model_name="Event",  # Передаем модель
            raw_data=data  # Передаем исходные данные
        )
        # event_data = {
        #     "town": event_town,
        #     "event": event_name,
        #     "date": event_date,
        #     "time": event_time,
        #     "link_to_source": link_to_source,
        #     "price": price,
        #     "place": place,
        #     "action": "активно",  # Статус по умолчанию
        # }
        # Формируем фильтры для поиска существующей записи
        filters = get_filters_for_model(
            model_name="Event",  # Указываем модель
            row_data=event_data  # Передаем данные
        )

        # # Формируем фильтры для поиска существующей записи
        # filters = {
        #     "town": event_town,
        #     "event": event_name,
        #     "date": event_date,
        # }

        # Пытаемся добавить город

        # Вызываем универсальную функцию
        result = await add_or_update_record(
            model_name="Event",# Передаем модель
            filters=filters,# Передаем фильтры
            data=event_data,# Передаем данные
        )
        
        if result:
            await message.answer(f"Событие '{event}' в городе {town} успешно добавлено!")
            # Отправляем уведомления всем пользователям

            # message_text = await create_new_message_text(
            #     town=town,
            #     event=event,
            #     description=description,
            #     event_date=event_date,
            #     time=time,
            #     link_to_source=link_to_source,
            #     price=price,
            #     place=place,
            # )            
            # await send_messages_to_users_all(
            #         bot=bot,
            #         message=message_text, 
            #         model_name="Event"
            #     )           

        else:
            await message.answer(f"Событие '{event} не получилось добавить'.")
    
    except Exception as e:
        # Обрабатываем возможные ошибки
        await message.answer(f"Произошла ошибка при добавлении события: {e}")
    
    finally:
        # Очищаем состояние
        await state.clear()

# @router.message(RemoveEvent.waiting_for_event_id)
# async def handle_remove_event_id(message: Message, state: FSMContext):
#     """
#     Обрабатывает ввод ID события для удаления.
#     """
#     event_id = int(message.text)

#     try:
#         # Пытаемся добавить город
#         # result = await remove_event(
#         #     event_id
#         # )
#         result = await remove_record(
#             model_name="Event",  # Передаем модель
#             filters={"id": event_id},  # Фильтр по названию города
#         )
        
#         if result:
#             await message.answer(f"Событие '{event_id}' успешно удалено!")
#         else:
#             await message.answer(f"Событие '{event_id} не получилось удалить'.")
    
#     except Exception as e:
#         # Обрабатываем возможные ошибки
#         await message.answer(f"Произошла ошибка при удалении события: {e}")
    
#     finally:
#         # Очищаем состояние
#         await state.clear()

@router.message(RemoveModel.waiting_for_id)
async def handle_remove_by_id(message: Message, state: FSMContext):
    """
    Обрабатывает ввод ID для удаления записи.
    """
    try:
        # Получаем ID из сообщения
        record_id = int(message.text)  # Преобразуем в число
        data = await state.get_data()
        model_name = data["model_name"]  # Получаем название модели из состояния

        # Удаляем запись по ID
        success = await remove_record(
            model_name=model_name,  # Передаем модель
            filters={"id": record_id},  # Фильтр по названию города
            )

        if success:
            await message.answer(f"Запись с ID {record_id} успешно удалена из модели {model_name}.")
        else:
            await message.answer(f"Не удалось удалить запись с ID {record_id}.")

    except ValueError:
        # Если введено не число
        await message.answer("Пожалуйста, введите корректный ID (число).")
    except Exception as e:
        # Обработка других ошибок
        await message.answer(f"Произошла ошибка при удалении: {e}")
    finally:
        # Очищаем состояние
        await state.clear()

@router.callback_query(F.data == "update_action")
async def start_update_event_action(callback: CallbackQuery, state: FSMContext):

    """
    Начинает процесс обновления статуса события.
    """
    if callback.from_user.id in admins:
        await callback.message.answer("Введите ID события:")
        await state.set_state(UpdateEventAction.waiting_for_event_id)
    else:
        await callback.message.answer("Извините, у Вас нет прав доступа")

@router.message(UpdateEventAction.waiting_for_event_id)
async def handle_event_id(message: Message, state: FSMContext):
    """
    Обрабатывает ввод ID события и предлагает выбрать новый статус.
    """
    try:
        event_id = int(message.text)
        await state.update_data(event_id=event_id)

        # Создаем клавиатуру с кнопками
        keyboard = kb.create_action_keyboard()

        # Отправляем сообщение с клавиатурой
        await message.answer(
            "Выберите новый статус:",
            reply_markup=keyboard,
        )
        await state.set_state(UpdateEventAction.waiting_for_new_action)
    except ValueError:
        await message.answer("Пожалуйста, введите корректный ID события (число).")

@router.message(UpdateEventAction.waiting_for_new_action)
async def handle_new_action(message: Message, state: FSMContext, bot: Bot):
    """
    Обрабатывает выбор нового статуса и обновляет событие.
    """
    new_action = message.text
    data = await state.get_data()
    event_id = data["event_id"]
    

    # Проверяем, что выбранный статус допустим
    if new_action not in ["активно", "отменено", "перенесено"]:
        await message.answer("Пожалуйста, выберите статус из предложенных кнопок.")
        return

    try:
        # Получаем полные данные о мероприятии
        event = await get_event_by_id(event_id)
        

        if event:
            # Обновляем статус события
            success = await update_event_action(event_id=event_id, new_action=new_action)
            
            if success:
                # Отправляем уведомления всем пользователям
                message_text = await create_new_message_text(
                    town=event.town,
                    event=event.event,
                    description=event.description,
                    event_date=event.event_date,
                    time=event.time,
                    link_to_source=event.link_to_source,
                    price=event.price,
                    place=event.place,
                    new_action=new_action,            
                )            
                await send_messages_to_users_all(
                            bot=bot,
                            message=message_text,
                            model_name="Event"
                        )  
                

                await message.answer(
                    f"Статус события '{event.event}' в городе {event.town} успешно обновлен на '{new_action}'.",
                    reply_markup=None,  # Убираем клавиатуру
                )
            else:
                await message.answer(
                    f"Не удалось обновить статус события с ID {event_id}.",
                    reply_markup=None,  # Убираем клавиатуру
                )
        else:
            await message.answer(
                f"Мероприятие с ID {event_id} не найдено.",
                reply_markup=None,  # Убираем клавиатуру
            )
    
    except Exception as e:
        await message.answer(
            f"Произошла ошибка при обновлении статуса: {e}",
            reply_markup=None,  # Убираем клавиатуру
        )
    
    finally:
        await state.clear()


@router.callback_query(F.data == "ban_action")
async def handle_ban_action(callback: CallbackQuery, state: FSMContext):
    """
    Обрабатывает нажатие кнопки "Забанить или разбанить пользователя".
    """
    if callback.from_user.id in admins:
        await callback.message.answer("Введите tg_ID пользователя:")
        await state.set_state(BanUser.waiting_for_user_id)
        await callback.answer()
    else:
        await callback.message.answer("Извините, у Вас нет прав доступа")

@router.message(BanUser.waiting_for_user_id)
async def handle_user_id(message: Message, state: FSMContext):
    """
    Обрабатывает ввод tg_ID пользователя.
    """
    user_id = message.text

    try:
        user_id = int(user_id)  # Преобразуем введенный текст в число
    except ValueError:
        await message.answer("Пожалуйста, введите корректный ID (число).")
        return

    # Проверяем, существует ли пользователь с таким ID
    user = await get_user_by_tg_id(user_id)
    if user:
        # Сохраняем ID пользователя в состоянии
        await state.update_data(user_id=user_id)

        # Создаем клавиатуру с кнопками "Забанить" и "Разбанить"
        keyboard = kb.create_ban_unban_keyboard()

        await message.answer(
            f"Пользователь {user.tg_name} (tg_ID: {user_id}) найден. Выберите действие:",
                reply_markup=keyboard,
        )
        await state.set_state(BanUser.waiting_for_ban_action)
    else:
        await message.answer(f"Пользователь с ID {user_id} не найден.")
        await state.clear()

@router.callback_query(BanUser.waiting_for_ban_action, F.data.in_(["ban_user", "unban_user"]))
async def handle_ban_choice(callback: CallbackQuery, state: FSMContext):
    """
    Обрабатывает выбор действия (бан/разбан).
    """
    data = await state.get_data()
    user_id = data["user_id"]  # Получаем ID пользователя из состояния

    if callback.data == "ban_user":
        success = await update_user_ban_status(user_id, ban_status=True)
        action_text = "забанен"
    else:
        success = await update_user_ban_status(user_id, ban_status=False)
        action_text = "разбанен"

    if success:
        await callback.message.answer(
            f"Пользователь с ID {user_id} успешно {action_text}."
        )
    else:
        await callback.message.answer(f"Пользователь с ID {user_id} не найден.")

    await state.clear()  # Очищаем состояние
    await callback.answer()            

# Обрабатываем выбор города после того как зашли в город
@router.callback_query(F.data.startswith("town_view_"))
async def handle_town_selection(callback: CallbackQuery):
    """
    Обрабатывает выбор города при зайти в город.
    """
    # Извлекаем название города из callback_data
    town = callback.data.split("_")[2]

    # Получаем количество записей для всех моделей
    record_counts = await get_record_counts_for_models(combined_model_names_for_users, town)

    # Создаем клавиатуру с передачей количества записей
    keyboard = kb.create_models_for_users_keyboard(town, combined_model_names_for_users, record_counts)

    # Отправляем сообщение с клавиатурой
    await callback.message.answer(f"Выберите раздел для города {town}:", reply_markup=keyboard)

    # Подтверждаем обработку callback
    await callback.answer()

@router.callback_query(F.data.startswith("model_"))
async def handle_model_selection(callback: CallbackQuery):
    """
    Обрабатывает выбор модели (например, "Мероприятия") после ЗАйти в город, для юзеров то есть.
    """
    # Извлекаем тип модели и город из callback_data
    model_name, town = callback.data.split("_")[1], callback.data.split("_")[2]

    if model_name == "Event":
        # Создаем клавиатуру для мероприятий
        keyboard = kb.create_event_periods_keyboard(town)

        # Отправляем сообщение с клавиатурой
        await callback.message.answer("Выберите период:", reply_markup=keyboard)
    elif model_name == "BusSchedule":
        # Создаем клавиатуру для мероприятий
        keyboard = kb.create_section_keyboard(prefix='getBus', section=sectionBus, town=town, model_name=model_name)

        # Отправляем сообщение с клавиатурой
        await callback.message.answer("Выберите раздел:", reply_markup=keyboard)
    else:
        # Получаем словарь для модели
        model_dict = all_sections.get(model_name) 

        if not model_dict or not isinstance(model_dict, dict):  # Проверяем, что это словарь
            await callback.message.answer("Ошибка: словарь не найден или не является словарем.")
            return
        
        # Подсчитываем количество записей для каждой секции
        counts = {}
        for section_key in model_dict.keys():
            # Получаем русское название раздела
            section_name_ru = get_section_name_ru(section_key, model_dict)

            # Формируем фильтры
            filters = {
                "town": town,
                "section": section_name_ru
            }

            # Получаем количество записей для раздела
            record_count = await get_record_count_analog(model_name, filters)
            counts[section_key] = record_count

        
        # Получаем клавиатуру из отдельной функции
        keyboard = kb.create_section_keyboard('getAnalog', model_dict, town, model_name, counts)
        
        await callback.message.answer("Выберите раздел:", reply_markup=keyboard)
   
    # Подтверждаем обработку callback
    await callback.answer()


@router.callback_query(F.data.startswith("getAnalog_"))
async def handle_getAnalog_section(callback: CallbackQuery):
    """
    Обрабатывает выбор раздела для показа подразделов аналогичных моделей.
    Формирует клавиатуру с кнопками, где текст кнопки — это имя + описание,
    а callback_data — это модель_id.
    """
    # Извлекаем данные из callback_data
    data = callback.data.split("_")
    section_name = data[1]  # Название раздела (например, "ITDevelopers")
    town_name = data[2]     # Название города (например, "Грязи")
    model_name = data[3]    # Название модели (например, "Services")

    # Получаем русское название раздела
    section_name_ru = get_section_name_ru(section_name, all_sections.get(model_name, {}))

    # Формируем фильтры
    filters = {"section": section_name_ru}

    # Получаем данные из базы данных для модели с фильтрами по городу и разделу
    records = await get_all_in_town(model_name, town_name, filters)

    if not records:
        await callback.message.answer("Данных не найдено.")
        await callback.answer()
        return
    
    # Создаем клавиатуру
    keyboard = kb.create_list_in_analog_models_keyboard(records, model_name)

    # Отправляем сообщение с клавиатурой
    await callback.message.answer("Выберите запись:", reply_markup=keyboard)

    # Подтверждаем обработку callback
    await callback.answer()

@router.callback_query(F.data.startswith("details_"))
async def handle_details(callback: CallbackQuery):
    """
    Обрабатывает нажатие на кнопку с callback_data в формате "details_модель_id".
    Показывает полную информацию о записи.
    """
    # Извлекаем данные из callback_data
    data = callback.data.split("_")
    model_name = data[1]  # Название модели (например, "Services")
    record_id = int(data[2])  # ID записи

    # Получаем запись по ID
    record = await get_record_by_id(model_name, record_id)
    if not record:
        await callback.message.answer("Запись не найдена.")
        await callback.answer()
        return

    # Формируем сообщение с полной информацией
    message_text = get_full_info_message(record)

    filters={
        "offerId": record_id
    }

    check_Record=await get_all_records(
        model_name="Record", filters=filters
    )

    if check_Record:
        keyboard = kb.booking_kb(model_name=model_name, record_id=record_id)
        # Отправляем сообщение
        await callback.message.answer(
            message_text, reply_markup=keyboard,parse_mode="HTML"
            )
    else:
        # Отправляем сообщение
        await callback.message.answer(message_text, parse_mode="HTML")

    # Проверяем, заканчивается ли модель на "Reserv" или отсутствует в all_sections
    if model_name.endswith("Reserv") or model_name not in all_sections:
        # Создаем клавиатуру с кнопками "Согласовать" и "Отменить"
        keyboard = kb.create_approval_keyboard(model_name, record_id)
        await callback.message.answer("Выберите действие:", reply_markup=keyboard)

    # Подтверждаем обработку callback
    await callback.answer()


@router.callback_query(F.data.startswith("getBus_"))
async def handle_getBus_section(callback: CallbackQuery):
    """
    Обрабатывает выбор раздела для показа расписания пользователю.
    """
    # Извлекаем данные из callback_data
    data = callback.data.split("_")
    section_name = data[1]
    town_name = data[2]
    model_name = data[3]

    # Преобразуем название раздела в русское название
    section_name_ru = get_section_name_ru(section_name, sectionBus)

    # Формируем фильтры
    filters = {"section": section_name_ru}

    # Получаем расписание
    shedules = await get_all_in_town(model_name, town_name, filters)

    if shedules:
        # Формируем сообщение с помощью новой функции
        message = create_schedule_message(shedules, town_name, section_name_ru)
        
        if message is None:
            # Если сообщение превышает лимит, предлагаем скачать Excel-файл
            logger.info("Превышает лимит")
            await callback.message.answer(
                "Сообщение слишком длинное. Расписание можете посмотреть на сайте: https://gryazy.ru/raspisanie_avtobus.html или можете скачать Excel-файл здесь",
                reply_markup=kb.create_download_excel_keyboard(town_name, model_name, section_name)
            )
            
            
        else:
            # Если сообщение не превышает лимит, отправляем его
            logger.info("НЕ превышает лимит")
            await callback.message.answer(message, parse_mode="Markdown")
    else:
        await callback.message.answer("Данных не найдено.")

    # Подтверждаем обработку callback
    await callback.answer()

@router.callback_query(F.data.startswith("event_today_"))
async def handle_event_today(callback: CallbackQuery):
    """
    Обрабатывает выбор мероприятий на сегодня.
    """
    # Извлекаем название города из callback_data
    town = callback.data.split("_")[2]

    date_day_today = get_today_date_dmy()

    # Получаем мероприятия на сегодня
    events = await get_events_by_date_and_town(event_date=date_day_today, town=town)

    if events:
        message = format_data_message(events, "Мероприятия на сегодня")
        if message =="Сообщение слишком длинное. Хотите скачать Excel-файл со всеми записями?":
            for event in events:
                message_text = format_data_one_message(event)
                # Отправляем сообщение
                await callback.message.answer(message_text)
        else:
            message_text = message
            await callback.message.answer(message_text)
    else:
        message_text = "На сегодня мероприятий нет."
        # Отправляем сообщение
        await callback.message.answer(message_text)

    # Подтверждаем обработку callback
    await callback.answer()


@router.callback_query(F.data.startswith("event_tomorrow_"))
async def handle_event_tomorrow(callback: CallbackQuery):
    """
    Обрабатывает выбор мероприятий на завтра.
    """
    # Извлекаем название города из callback_data
    town = callback.data.split("_")[2]

    date_day_tomorrow=get_tomorrow_date()

    # Получаем мероприятия на завтра
    events = await get_events_by_date_and_town(event_date=date_day_tomorrow, town=town)

    if events:
        message = format_data_message(events, "Мероприятия на завтра")
        if message =="Сообщение слишком длинное. Хотите скачать Excel-файл со всеми записями?":
            for event in events:
                message_text = format_data_one_message(event)
                # Отправляем сообщение
                await callback.message.answer(message_text)
        else:
            message_text = message
            await callback.message.answer(message_text)
    else:
        message_text = "На завтра мероприятий нет."
        # Отправляем сообщение
        await callback.message.answer(message_text)

    # Подтверждаем обработку callback
    await callback.answer()




@router.callback_query(F.data.startswith("event_future_"))
async def handle_event_future(callback: CallbackQuery):
    """
    Обрабатывает выбор будущих мероприятий.
    """
    # Извлекаем название города из callback_data
    town = callback.data.split("_")[2]

    date_day_today=get_today_date_dmy()

    # Получаем будущие мероприятия
    events = await get_events_future_by_town(town=town, event_date=date_day_today)
    
    if events:
        # Сортируем мероприятия по дате
        sorted_events = sorted(
            events,
            key=lambda x: (x['event_date'])  # Используем ключи словаря
        )
        try:
            message = format_data_message(sorted_events, "Будущие мероприятия")
            if message =="Сообщение слишком длинное. Хотите скачать Excel-файл со всеми записями?":
                keyboard = kb.create_filters_events_keyboard(town=town)
                await callback.message.answer("Мероприятий слишком много, лучше выбрать фильтр:", reply_markup=keyboard)
            else:
                await callback.message.answer(message)
        except Exception as e:
            logger.error(f"Ошибка при создании сообщения: {e}")
    else:
        message_text = "Будущих мероприятий нет."
        await callback.message.answer(message_text)

    # Подтверждаем обработку callback
    await callback.answer()


@router.callback_query(F.data.startswith("Event_all_"))
async def handle_all_events(callback: CallbackQuery):
    """
    Обрабатывает нажатие на кнопку "Все мероприятия".
    """
    # Извлекаем название города из callback_data
    town = callback.data.split("_")[2]

    # Извлекаем название модели из callback_data
    model_name = callback.data.split("_")[0]  # Получаем "Event"

    # Получаем все мероприятия
    events = await get_all_in_town(model_name, town)

    if events:
        # Преобразуем объекты SQLAlchemy в словари
        # events_dicts = [
        #     {
        #         "town": event.town,
        #         "event": event.event,
        #         "date": event.date,
        #         "time": event.time,
        #         "place": event.place,
        #         "price": event.price,
        #         "link_to_source": event.link_to_source,
        #         "action": event.action,
        #     }
        #     for event in events
        # ]
        # Преобразуем объекты SQLAlchemy в словари
        events_dicts = convert_orm_to_dict(events, model_name)

        # Сортируем мероприятия по дате и времени
        sorted_events = sorted(
            events_dicts,
            key=lambda x: (x["event_date"])  # Сначала по дате, затем по времени
        )

        # Формируем сообщение с помощью format_data_message
        message = format_data_message(sorted_events, title=f"Все мероприятия в городе {town}")

        if message =="Сообщение слишком длинное. Хотите скачать Excel-файл со всеми записями?":
            # Если сообщение превышает лимит, предлагаем скачать Excel-файл
            logger.info("Превышает лимит")
            await callback.message.answer(
                "Сообщение слишком длинное. Хотите скачать Excel-файл со всеми мероприятиями?",
                reply_markup=kb.create_download_excel_keyboard(town, model_name)
            )
        else:
            # Если сообщение не превышает лимит, отправляем его
            logger.info("НЕ превышает лимит")
            await callback.message.answer(message, parse_mode="Markdown")
    else:
        await callback.message.answer("Мероприятий не найдено.")

    # Подтверждаем обработку callback
    await callback.answer()
    



# @router.callback_query(F.data == "get_")
@router.callback_query(F.data.startswith("get_"))
async def handle_all_events(callback: CallbackQuery):
    """
    Обрабатывает нажатие на кнопку "Все мероприятия".
    Если сообщение слишком длинное, предлагает скачать Excel-файл.
    """
    # Извлекаем название модели из callback_data
    model_name = callback.data.split("_")[1]  # Получаем "Event"

    if callback.from_user.id in admins:
        # Получаем все записи для модели
        records = await get_all_records(model_name)
        model = get_model_by_name(model_name)

        if records:
            # Формируем сообщение
            message = await generate_message_from_model(model=model, records=records)
            if message == "Сообщение слишком длинное. Хотите скачать Excel-файл со всеми записями?":
                # Предлагаем скачать Excel-файл
                await callback.message.answer(
                    message,
                    reply_markup=kb.create_download_excel_keyboard("all", model_name)
                )
            else:
                # Отправляем сообщение
                await callback.message.answer(message, parse_mode="Markdown")
        else:
            await callback.message.answer("Записей не найдено.")

        # Подтверждаем обработку callback
        await callback.answer()
    else:
        await callback.message.answer("Извините, у Вас нет прав доступа")

async def send_excel_file(callback: CallbackQuery, filename: str, caption: str):
    """
    Отправляет Excel-файл пользователю и удаляет временный файл.

    :param callback: Объект CallbackQuery.
    :param filename: Имя файла.
    :param caption: Подпись к файлу.
    """
    file = FSInputFile(filename)
    await callback.message.answer_document(file, caption=caption)
    os.remove(filename)

@router.callback_query(F.data.startswith("download_excel_"))
async def download_excel(callback: CallbackQuery):
    """
    Формирует и отправляет Excel-файл с данными (пользователи, мероприятия и т.д.).
    builder.button(text="Скачать Excel", callback_data=f"download_excel_{town}_{model_name}_{section}")
    """

    await callback.message.answer(
        "Файл подготавливается, ожидайте")
    
    # Разделяем callback_data по символу "_"
    parts = callback.data.split("_")

    # Проверяем, что parts содержит достаточно элементов
    if len(parts) < 4:
        await callback.message.answer("Ошибка: некорректный формат callback_data.")
        await callback.answer()
        return

    town_or_all = parts[2]  # Город или "all"
    model_name = parts[3]   # Имя модели (например, "User", "Event", "BusSchedule")

    # Проверяем, есть ли section_name в callback_data
    section_name = parts[4] if len(parts) > 4 else None  # Имя раздела (если есть)

    # section_name = parts[4]   # Имя модели
    section_name_ru=get_section_name_ru(section_name, sectionBus)
    # section_name_ru=sectionBus.get(section_name, "Неизвестный раздел")

    # Получаем данные в зависимости от модели
    if town_or_all == "all":
        records = await get_all_records(model_name)
    else:
        if section_name:
            filters = {"section": section_name_ru}
            records = await get_all_in_town(model_name, town_or_all, filters)
        else:
            records = await get_all_in_town(model_name, town_or_all)

    if records:
        # Подготавливаем данные для Excel с использованием prepare_model_data
        data = []
        headers = None

        

        for record in records:
            # Преобразуем запись в словарь (если это ORM-объект)
            raw_data = record.__dict__ if hasattr(record, "__dict__") else record

            # Подготавливаем данные для модели
            prepared_data = prepare_model_data(model_name, raw_data)

            # Формируем заголовки, если они еще не заданы
            if headers is None:
                headers = list(prepared_data.keys())

            # Добавляем строку данных
            data.append(list(prepared_data.values()))

        # Формируем имя файла
        filename = f"{model_name.lower()}_{town_or_all}.xlsx"

        # Отправляем Excel-файл
        await create_and_send_excel(
            callback,
            data,
            headers,
            filename,
            model_name,  # Название листа (например, "User", "Event", "BusSchedule")
            f"Данные для {model_name} ({town_or_all})"  # Заголовок
        )
    else:
        await callback.message.answer(f"Данные для модели {model_name} не найдены.")
        await callback.answer()


   
async def create_and_send_excel(callback: CallbackQuery, data: list, headers: list, filename: str, sheet_title: str, caption: str):
    """
    Создает Excel-файл и отправляет его пользователю.

    :param callback: Объект CallbackQuery.
    :param data: Список данных.
    :param headers: Заголовки столбцов.
    :param filename: Имя файла.
    :param sheet_title: Название листа.
    :param caption: Подпись к файлу.
    """
    try:
        # Создаем Excel-файл
        await create_excel_file(data, headers, filename, sheet_title)
        # Отправляем файл пользователю
        await send_excel_file(callback, filename, caption)
        await callback.message.answer(
            "Файл готов, можете скачать его")
    except Exception as e:
        logger.error(f"Ошибка при создании или отправке Excel-файла: {e}")
        await callback.message.answer("Произошла ошибка при создании файла.")
    finally:
        # Подтверждаем обработку callback
        await callback.answer()

@router.callback_query(F.data == "view_town_all_with_id")
async def handle_all_events(callback: CallbackQuery):
    """
    Обрабатывает нажатие на кнопку "Все мероприятия".
    """
    
    if callback.from_user.id in admins:
        # Получаем все мероприятия
        towns = await get_all_records("Town")

        if towns:
            # Здесь не стала в общую функцию вгонять, так как нужен еще id

            
            # Формируем сообщение с мероприятиями
            message = f"Вся информация по городам:\n\n"
            for town in towns:
                
                message += (
                    f"**ID:** {town.id}\n"
                    f"🏙️ **Город:** {town.town}\n"
                    f"🎭 **Грейд:** {town.grade}\n"
                    "-------------------------\n\n"
                )
        else:
            message = "Городов не найдено."

        # Отправляем сообщение
        await callback.message.answer(message, parse_mode="Markdown")

        # Подтверждаем обработку callback
        await callback.answer()
    else:
        await callback.message.answer("Извините, у Вас нет прав доступа")

@router.callback_query(UploadData.waiting_for_model, F.data.startswith("upload_"))
async def handle_upload_model(callback: CallbackQuery, state: FSMContext):
    """
    Обрабатывает выбор модели для загрузки данных.
    """
    model_name = callback.data.replace("upload_", "")  # Извлекаем название модели
    await state.update_data(model_name=model_name)  # Сохраняем выбранную модель в состоянии
    await callback.message.answer(f"Вы выбрали модель '{combined_model_names_for_admin[model_name]}'. Пожалуйста, загрузите Excel-файл с данными.")
    await state.set_state(UploadData.waiting_for_file)
    await callback.answer()


@router.message(UploadData.waiting_for_file, F.document)
async def handle_upload_file(message: Message, state: FSMContext, bot: Bot):
    """
    Обрабатывает загрузку Excel-файла и добавляет данные в базу данных.
    """
    # Получаем данные из состояния
    data = await state.get_data()
    model_name = data["model_name"]

    # Скачиваем файл
    file_id = message.document.file_id
    file = await bot.get_file(file_id)
    file_path = file.file_path

    # Скачиваем файл в память
    downloaded_file = await bot.download_file(file_path)
    file_bytes = downloaded_file.read()

    # Счетчик добавленных записей
    added_count = 0

    try:
        # Читаем Excel-файл с помощью openpyxl
        workbook = load_workbook(filename=BytesIO(file_bytes))
        sheet = workbook.active  # Получаем активный лист

        # Получаем заголовки столбцов (первая строка)
        headers = [cell.value for cell in sheet[1]]

        # Получаем колонки модели
        model_columns = get_model_columns(model_name)

        # Обрабатываем строки Excel-файла
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовки
            row_data = dict(zip(headers, row))  # Создаем словарь из заголовков и значений

            # Преобразуем дату и время в строки, если это необходимо
            # if "date" in row_data and isinstance(row_data["date"], datetime):
            #     row_data["date"] = row_data["date"].strftime("%Y-%m-%d")  # Формат: YYYY-MM-DD
            if "time" in row_data and isinstance(row_data["time"], time):
                row_data["time"] = row_data["time"].strftime("%H:%M")  # Формат: HH:MM
            if "time_start" in row_data and isinstance(row_data["time_start"], time):
                row_data["time_start"] = row_data["time_start"].strftime("%H:%M")  # Формат: HH:MM
            if "time_finish" in row_data and isinstance(row_data["time_finish"], time):
                row_data["time_finish"] = row_data["time_finish"].strftime("%H:%M") # Формат: HH:MM
            # Преобразуем price в строку, если это необходимо
            if "price" in row_data and isinstance(row_data["price"], (int, float)):
                row_data["price"] = str(row_data["price"])  # Преобразуем в строку
            if "number" in row_data and isinstance(row_data["number"], (int, float)):
                row_data["number"] = str(row_data["number"])  # Преобразуем в строку
          
            # Фильтруем row_data, оставляя только те ключи, которые есть в параметрах функции
            filtered_data = {k: v for k, v in row_data.items() if k in model_columns}

            # Получаем фильтры для модели
            filters = get_filters_for_model(model_name, row_data)

            # Вызываем универсальную функцию для добавления или обновления записи
            success = await add_or_update_record(
                model_name=model_name,
                filters=filters,
                data=filtered_data,
            )

            if success:
                added_count += 1  # Увеличиваем счетчик добавленных записей
                logger.info(f'Данные успешно добавлены в базу')

        # Формируем сообщение о количестве добавленных записей
        # logger.info(f'Формируем сообщение о добавленных записях')
        # if added_count > 0:
        #     if model_name == "Event":
        #         # Список для хранения всех записей
        #         all_records = []
        #         all_records.append(filtered_data)  # Добавляем запись в список
        #         message_text = format_data_message(events=all_records)
        #         if message_text == "Сообщение слишком длинное. Хотите скачать Excel-файл со всеми записями?":
        #             for event in filtered_data:
        #                 message_text = format_data_one_message(event=event)
        #                 await send_messages_to_users_all(
        #                     bot=bot,
        #                     message=message_text,
        #                     model_name=model_name,
        #                 )
        #         else:
        #             await send_messages_to_users_all(
        #                 bot=bot,
        #                 message=message_text,
        #                 model_name=model_name,
        #             )   
        #     else:
        #         message_text = create_upload_message(model_name, added_count)
        #         await send_messages_to_users_all(
        #                 bot=bot,
        #                 message=message_text,
        #                 model_name=model_name,
        #             )
      
        await message.answer(f"Данные для модели '{model_name}' успешно загружены! Добавлено {added_count} записей.")

    except Exception as e:
        logger.error(f"Ошибка при загрузке данных: {e}")
        await message.answer(f"Произошла ошибка при загрузке данных: {e}")

    # Очищаем состояние
    await state.clear()


@router.callback_query(F.data == "update_grade")
async def handle_update_grade(callback: CallbackQuery, state: FSMContext):
    """
    Обрабатывает нажатие на кнопку "Изменить грейд показа".
    """
    if callback.from_user.id in admins:
        # Получаем список моделей
        models = await get_models(combined_model_names_for_admin)

        # Создаем клавиатуру с моделями
        keyboard = kb.create_models_keyboard(models, combined_model_names_for_admin, "update_grade")
        await callback.message.answer("Выберите модель для изменения грейда:", reply_markup=keyboard)
        await state.set_state(UpdateGrade.waiting_for_model)
    else:
        await callback.message.answer("Извините, у Вас нет прав доступа")
    await callback.answer()

@router.callback_query(UpdateGrade.waiting_for_model, F.data.startswith("update_grade_"))
async def handle_update_grade_model(callback: CallbackQuery, state: FSMContext):
    """
    Обрабатывает выбор модели для изменения грейда.
    """
    model_name = callback.data.replace("update_grade_", "")  # Извлекаем название модели
    await state.update_data(model_name=model_name)  # Сохраняем выбранную модель в состоянии
    await callback.message.answer(f"Вы выбрали модель '{combined_model_names_for_admin[model_name]}'. Введите ID строки, которую хотите изменить:")
    await state.set_state(UpdateGrade.waiting_for_id)
    await callback.answer()

@router.message(UpdateGrade.waiting_for_id)
async def handle_update_grade_id(message: Message, state: FSMContext):
    """
    Обрабатывает ввод ID строки для изменения грейда.
    """
    try:
        row_id = int(message.text)  # Преобразуем введенный текст в число
        await state.update_data(row_id=row_id)  # Сохраняем ID строки в состоянии
        await message.answer("Введите новое значение грейда:")
        await state.set_state(UpdateGrade.waiting_for_grade)
    except ValueError:
        await message.answer("Пожалуйста, введите корректный ID (число).")

@router.message(UpdateGrade.waiting_for_grade)
async def handle_update_grade_value(message: Message, state: FSMContext):
    """
    Обрабатывает ввод нового грейда и обновляет данные в базе.
    """
    try:
        new_grade = int(message.text)  # Преобразуем введенный текст в число
        
        data = await state.get_data()
        model_name = data["model_name"]
        row_id = int(data["row_id"])
        

        # Обновляем грейд в базе данных
        success=await update_grade(model_name, row_id, new_grade)

        if success:
            await message.answer(f"Грейд для строки с ID {row_id} успешно обновлен на {new_grade}.")
        else:
            await message.answer(f"Не удалось обновить грейд для строки с ID {row_id}.")

    except ValueError:
        await message.answer("Пожалуйста, введите корректное значение грейда (число).")
    finally:
        await state.clear()

@router.callback_query(F.data == "view_User_all")
async def handle_all_events(callback: CallbackQuery):
    """
    Обрабатывает нажатие на кнопку "Посмотреть всех юзеров".
    Если сообщение слишком длинное, предлагает скачать Excel-файл.
    """
    # Извлекаем название модели из callback_data
    model_name = callback.data.split("_")[1]  # Получаем "User"
    
    if callback.from_user.id in admins:
        # Получаем всех пользователей
        users = await get_all_records(model_name)

        if users:
            # Формируем сообщение с помощью новой функции
            message = create_user_message(users)
            
            if message is None:
                # Если сообщение превышает лимит, предлагаем скачать Excel-файл
                await callback.message.answer(
                    "Сообщение слишком длинное. Хотите скачать Excel-файл со всеми пользователями?",
                    reply_markup=kb.create_download_excel_keyboard("all", model_name)
                )
            else:
                 # Логируем сообщение перед отправкой
                logger.info(f'Сообщение для отправки: {message}')

                # Если сообщение не превышает лимит, отправляем его
                # await callback.message.answer(message, parse_mode="Markdown")
                await callback.message.answer(message, parse_mode="HTML")
        else:
            await callback.message.answer("Пользователи не найдены.")

        # Подтверждаем обработку callback
        await callback.answer()
    else:
        await callback.message.answer("Извините, у Вас нет прав доступа")


@router.message(AddBusSchedule.waiting_for_town)
async def handle_bus_schedule_town(message: Message, state: FSMContext):
    """
    Обрабатывает ввод города для расписания автобусов.
    """
    town=message.text
    await state.update_data(town=town)
    
    # Получаем клавиатуру из отдельной функции
    keyboard = kb.create_section_keyboard('addBuss', sectionBus, town)
    
    await message.answer("Выберите раздел:", reply_markup=keyboard)
    await state.set_state(AddBusSchedule.waiting_for_number_bus)

@router.callback_query(AddBusSchedule.waiting_for_number_bus)
async def handle_bus_schedule_section(callback: CallbackQuery, state: FSMContext):
    """
    Обрабатывает выбор раздела автобусного расписания через инлайн-кнопку.
    """
    # Получаем данные из callback_data
    selected_section_key = callback.data.split("_")[1]

    section_name_ru = get_section_name_ru(selected_section_key, sectionBus)

    
    # Сохраняем русское название раздела в состояние
    await state.update_data(section=section_name_ru)
    
    # Отвечаем на callback, чтобы убрать "часики" у кнопки
    await callback.answer()
    
    # Отправляем сообщение с выбранным разделом
    await callback.message.answer(f"Вы выбрали раздел: {section_name_ru}")
    
    # Переводим состояние на следующий шаг
    await callback.message.answer("Введите номер маршрута:")
    await state.set_state(AddBusSchedule.waiting_for_section_bus)


@router.callback_query(AddBusSchedule.waiting_for_section_bus)
async def handle_bus_schedule_section(message: Message, state: FSMContext):
    """
    Обрабатывает ввод номера марштруа
    """
    number=str(message.text)
    await state.update_data(number=number)
    
    
    # Отвечаем на callback, чтобы убрать "часики" у кнопки
    await message.answer()
    
    # Отправляем сообщение с выбранным разделом
    await message.answer(f"Вы выбрали раздел: {number}")
    
    # Переводим состояние на следующий шаг
    await message.answer("Введите место отправления:")
    await state.set_state(AddBusSchedule.waiting_for_start_place)

@router.message(AddBusSchedule.waiting_for_start_place)
async def handle_bus_schedule_start_place(message: Message, state: FSMContext):
    """
    Обрабатывает ввод места отправления.
    """
    await state.update_data(start_place=message.text)
    await message.answer("Введите место назначения:")
    await state.set_state(AddBusSchedule.waiting_for_finish_place)

@router.message(AddBusSchedule.waiting_for_finish_place)
async def handle_bus_schedule_finish_place(message: Message, state: FSMContext):
    """
    Обрабатывает ввод места назначения.
    """
    await state.update_data(finish_place=message.text)
    await message.answer("Введите время отправления (например, 08:00):")
    await state.set_state(AddBusSchedule.waiting_for_time_start)

@router.message(AddBusSchedule.waiting_for_time_start)
async def handle_bus_schedule_time_start(message: Message, state: FSMContext):
    """
    Обрабатывает ввод времени отправления.
    """
    await state.update_data(time_start=message.text)
    await message.answer("Введите время прибытия (например, 10:00), если не известно, поставьте '-':")
    await state.set_state(AddBusSchedule.waiting_for_time_finish)

@router.message(AddBusSchedule.waiting_for_time_finish)
async def handle_bus_schedule_time_finish(message: Message, state: FSMContext):
    """
    Обрабатывает ввод времени прибытия.
    """

    time_finish = message.text  
    
    await state.update_data(time_finish=time_finish)

    # await state.update_data(time_finish=message.text)
    await message.answer("Введите дни работы (например, Ежедневно или Пн-Пт):")
    await state.set_state(AddBusSchedule.waiting_for_days)

@router.message(AddBusSchedule.waiting_for_days)
async def handle_bus_schedule_days(message: Message, state: FSMContext):
    """
    Обрабатывает ввод дней работы.
    """
    await state.update_data(days=message.text)
    await message.answer('Введите ссылку на источник (если есть, если нет, поставьте "-"):')
    await state.set_state(AddBusSchedule.waiting_for_link_to_source)

@router.message(AddBusSchedule.waiting_for_link_to_source)
async def handle_bus_schedule_link_to_source(message: Message, state: FSMContext):
    """
    Обрабатывает ввод ссылки на источник.
    """
    link_to_source=message.text

    await state.update_data(link_to_source=link_to_source)
        
    # await state.update_data(link_to_source=message.text)

    # Получаем все данные из состояния
    data = await state.get_data()
    model_name = data["model_name"]
    

    # Формируем данные для добавления записи
    bus_schedule_data = prepare_model_data(model_name, data)
    # bus_schedule_data = {
    #     "town": data["town"],
    #     "start_place": data["start_place"],
    #     "finish_place": data["finish_place"],
    #     "time_start": data["time_start"],
    #     "time_finish": data["time_finish"],
    #     "days": data["days"],
    #     "link_to_source": data["link_to_source"],
    # }

    filters=get_filters_for_model(model_name, bus_schedule_data)

    # Вызываем универсальную функцию для добавления или обновления записи
    result = await add_or_update_record(
        model_name=model_name,  # Передаем модель
        filters=filters,  # Фильтры для поиска существующей записи
        data=bus_schedule_data,  # Передаем данные
    )

    if result:
        await message.answer("Запись успешно добавлена/обновлена!")
    else:
        await message.answer("Произошла ошибка при добавлении записи.")

    # Очищаем состояние
    await state.clear()

@router.message(AddAnalogModel.waiting_for_town)
async def handle_bus_schedule_town(message: Message, state: FSMContext):
    """
    Обрабатывает ввод города для аналогичных моделей.
    """
    town=message.text
    await state.update_data(town=town)

    #Получаем данные из состояния
    data = await state.get_data()
    model_name = data.get("model_name")  # Имя модели (например, "Food", "Services" и т.д.)

    # Получаем словарь для модели
    model_dict = all_sections.get(model_name) # Получаем словарь по имени модели

    if not model_dict or not isinstance(model_dict, dict):  # Проверяем, что это словарь
        await message.answer("Ошибка: словарь не найден или не является словарем.")
        return
    
    # Получаем клавиатуру из отдельной функции
    keyboard = kb.create_section_keyboard(prefix='addAnalog', section=model_dict,  town=town, model_name=model_name)
    
    await message.answer("Выберите раздел:", reply_markup=keyboard)
    await state.set_state(AddAnalogModel.waiting_for_option)

@router.callback_query(AddAnalogModel.waiting_for_option)
async def handle_option_selection(callback: CallbackQuery, state: FSMContext):
    """
    Обрабатывает выбор подраздела (опции) из инлайн-клавиатуры.
    """
    # Разделяем callback_data на части
    callback_data = callback.data.split('_')
    
    section_key = callback_data[1]  # Ключ раздела (например, "FromGryaziBusStation")
    # town = callback_data[2]  # Город,
    model_name = callback_data[3]  # Имя модели (например, "BusSchedule")

    # Получаем русское название раздела
    section_name_ru = get_section_name_ru(section_key, all_sections.get(model_name, {}))

    # Сохраняем данные в состояние
    await state.update_data(
        section=section_name_ru
    )

    # Запрашиваем наименование организации
    await callback.message.answer(f"Вы выбрали раздел: {section_name_ru}. Введите наименование организации по которому Вас знают люди (н-р, Долгополова Виктория):")
    await state.set_state(AddAnalogModel.waiting_for_name)

    # Подтверждаем обработку callback
    await callback.answer()

@router.message(AddAnalogModel.waiting_for_name)
async def handle_name(message: Message, state: FSMContext):
    await state.update_data(name=message.text)
    await message.answer("Введите краткое описание (н-р, Разработчик сайтов, ПО, чат-ботов, автоматизация рабочих процессов):")
    await state.set_state(AddAnalogModel.waiting_for_description_small)

@router.message(AddAnalogModel.waiting_for_description_small)
async def handle_description_small(message: Message, state: FSMContext):
    descriptionSmall = message.text
    if len(descriptionSmall) > 500:
        await message.answer("Краткое описание слишком длинное. Будет видно только 500 символов.")
    
    await state.update_data(descriptionSmall=descriptionSmall[:500])
    await message.answer("Введите полное описание (н-р, Реализовываю собственные full-stack проекты, которые помогают людям решать разные задачи, достигать поставленных целей):")
    await state.set_state(AddAnalogModel.waiting_for_description_full)

@router.message(AddAnalogModel.waiting_for_description_full)
async def handle_description_full(message: Message, state: FSMContext):
    descriptionFull = message.text
    if len(descriptionFull) > 1500:
        await message.answer("Полное описание слишком длинное. Будет видно только 1500 символов.")
    
    await state.update_data(descriptionFull=descriptionFull[:1500])
    await message.answer("Введите график работы (н-р, ПН-ПТ с 8:00 до 17:00):")
    await state.set_state(AddAnalogModel.waiting_for_schedule)

@router.message(AddAnalogModel.waiting_for_schedule)
async def handle_schedule(message: Message, state: FSMContext):
    await state.update_data(schedule=message.text)
    await message.answer("Введите координаты местонахождения (н-р, 55.7558, 37.6176), если не хотите, то поставьте -:")
    await state.set_state(AddAnalogModel.waiting_for_coordinates)

@router.message(AddAnalogModel.waiting_for_coordinates)
async def handle_coordinates(message: Message, state: FSMContext):
    await state.update_data(coordinates=message.text)
    await message.answer("Введите адрес (н-р, г.Грязи, ул.Семашко), если не хотите, поставьте -:")
    await state.set_state(AddAnalogModel.waiting_for_address)

@router.message(AddAnalogModel.waiting_for_address)
async def handle_address(message: Message, state: FSMContext):
    await state.update_data(address=message.text)
    await message.answer("Введите телефон в формате(н-р, +7(904)-284-89-69):")
    await state.set_state(AddAnalogModel.waiting_for_phone)

@router.message(AddAnalogModel.waiting_for_phone)
async def handle_phone(message: Message, state: FSMContext):
    await state.update_data(phone=message.text)
    await message.answer("Введите сайт (н-р, https://dolgopolovav.ru), можете страницу в ВК:")
    await state.set_state(AddAnalogModel.waiting_for_website)

@router.message(AddAnalogModel.waiting_for_website)
async def handle_website(message: Message, state: FSMContext):
    await state.update_data(website=message.text)
    await message.answer("Напишите Ваше ФИО (н-р, Долгополова Виктория Викторовна)")
    await state.set_state(AddAnalogModel.waiting_for_nameUser)

@router.message(AddAnalogModel.waiting_for_nameUser)
async def handle_nameUser(message: Message, state: FSMContext):
    """
    Обрабатывает ввод имени пользователя и формирует сообщение для подтверждения.
    """
    # Обновляем состояние с введенным именем пользователя и его Telegram ID
    await state.update_data(nameUser=message.text)
    tgId = message.from_user.id
    await state.update_data(tgId=tgId)

    # Получаем все данные из состояния
    data = await state.get_data()

    # Формируем сообщение с полной информацией
    message_text = get_full_info_message_data(data)

    confirmation_keyboard=kb.confirmation_keyboard

    # Отправляем сообщение с полной информацией и клавиатурой
    await message.answer("Проверьте введенные данные:\n\n" + message_text, reply_markup=confirmation_keyboard)

    # Устанавливаем состояние ожидания подтверждения
    await state.set_state(AddAnalogModel.waiting_for_confirmation)


@router.message(AddAnalogModel.waiting_for_confirmation)
async def handle_confirmation(message: Message, state: FSMContext, bot: Bot):
    """
    Обрабатывает подтверждение или запрос на редактирование.
    """
    if message.text == "Верно":

        # Получаем все данные из состояния
        data = await state.get_data()


        # Удаляем лишние ключи, которые не относятся к модели(убарли модельнейм)
        model_fields = {
            "town", "section", "name", "descriptionSmall", "descriptionFull",
            "schedule", "coordinates", "address", "phone", "website",
            "nameUser", "tgId", "grade"
        }
        filtered_data = {key: value for key, value in data.items() if key in model_fields}

        # Обрезаем данные до допустимой длины
        if "descriptionSmall" in filtered_data:
            filtered_data["descriptionSmall"] = filtered_data["descriptionSmall"][:500]  # Обрезаем до 500 символов
        if "descriptionFull" in filtered_data:
            filtered_data["descriptionFull"] = filtered_data["descriptionFull"][:1500]  # Обрезаем до 1500 символов

        # Добавляем запись в базу данных
        model_name = data.get("model_name")
        town = data.get("town")
        section = data.get("section")

        filters = {"town": town, "section": section, "name": data.get("name")}
        if message.from_user.id in admins:
            success = await add_or_update_record(model_name, filters, filtered_data)

            if success:
                await message.answer("Запись успешно добавлена/обновлена!", reply_markup=ReplyKeyboardRemove())
            else:
                await message.answer("Произошла ошибка при добавлении записи.", reply_markup=ReplyKeyboardRemove())
        else:
            model_nameRes=f'{model_name}Reserv'
            success = await add_or_update_record(model_nameRes, filters, filtered_data)
            section_name_ru=get_section_name_ru(model_name, analog_model_names)

            if success:
                await message.answer(f'Спасибо! Ваше предложение отправлено на согласование! После согласования его можно будет увидеть в разделе "{section_name_ru}"', reply_markup=ReplyKeyboardRemove())
            else:
                await message.answer("Произошла ошибка при Отправки предложения на согласование.", reply_markup=ReplyKeyboardRemove())

            # Формируем сообщение с полной информацией
            message_text = get_full_info_message_data(data, model_name)
            # Отправляем сообщение админу с просьбой подтвердить
            user_id = admins[0]
            topic="Предложить добавить"
            status="предложение" # так как это предложение от пользователя
            user_name=message.from_user.username
            await send_message_to_user(
                                fromWhomUser=user_name,
                                bot=bot,
                                user_id=user_id,
                                topic=topic, 
                                message_text=message_text,
                                status=status,
                                )

        # Очищаем состояние
        await state.clear()

    elif message.text == "Редактировать":
        # Возвращаем пользователя на начальный шаг
        await message.answer("Хорошо, начнем заново. Введите город:", reply_markup=ReplyKeyboardRemove())
        await state.set_state(AddAnalogModel.waiting_for_town)
    else:
        # Если пользователь ввел что-то другое
        await message.answer("Пожалуйста, выберите 'Редактировать' или 'Верно'.", reply_markup=ReplyKeyboardRemove())


@router.callback_query(F.data == "send_messages")
async def handle_admin_send_messages(callback: CallbackQuery, state: FSMContext):
    """
    Обрабатывает нажатие на кнопку "Написать письмо" и "Написать админу\сообщить об ошибке".
    """
    
    # Отправляем сообщение с выбором получателя
    await callback.message.answer("Введите тему сообщения")

    # Устанавливаем состояние CHOOSE_RECIPIENT
    await state.set_state(SendMessageState.ENTER_TOPIC)

    # Подтверждаем обработку callback
    await callback.answer()
    

@router.message(SendMessageState.ENTER_TOPIC)
async def handle_enter_topic(message: Message, state: FSMContext):
    """
    Обрабатывает ввод темы сообщения.
    """

    # Сохраняем тему сообщения в состоянии
    await state.update_data(topic=message.text)

    if message.from_user.id in admins:
        # Создаем клавиатуру для выбора получателя
        keyboard = kb.send_admin_kb
    
        # Предлагаем выбрать получателя
        await message.answer("Кому вы хотите отправить сообщение?", reply_markup=keyboard)
        await state.set_state(SendMessageState.CHOOSE_RECIPIENT)
    else:
        # Для неадминистраторов сразу запрашиваем текст сообщения
        await message.answer("Введите текст сообщения:")
        await state.set_state(SendMessageState.ENTER_MESSAGE)

@router.callback_query(SendMessageState.CHOOSE_RECIPIENT)
async def handle_choose_recipient(callback: CallbackQuery, state: FSMContext):
    """
    Обрабатывает выбор получателя (всем или одному) если пусто, то этот шаг пропускается.
    """
    if callback.data == "send_to_all":
        # Если выбрано "всем", запрашиваем текст сообщения
        await callback.message.answer("Введите текст сообщения:")
        await state.set_state(SendMessageState.ENTER_MESSAGE)
    elif callback.data == "send_to_one":
        # # Если выбрано "одному", запрашиваем ID пользователя
        await callback.message.answer("Введите ID пользователя:")
        await state.set_state(SendMessageState.ENTER_USER_ID)
        

    # Подтверждаем обработку callback
    await callback.answer()


@router.message(SendMessageState.ENTER_USER_ID)
async def handle_enter_user_id(message: Message, state: FSMContext):
    """
    Обрабатывает ввод ID пользователя для отправки сообщения одному пользователю.
    """
    if message.from_user.id in admins:
        # Получаем ID пользователя
        user_id = message.text

        # Проверяем, что ID является числом
        if not user_id.isdigit():
            await message.answer("ID пользователя должен быть числом. Попробуйте снова.")
            return
        
        # Преобразуем user_id в int
        user_id = int(user_id)

        # Проверяем, существует ли пользователь с таким ID
        user = await get_user_by_tg_id(user_id)
        if user:
            # Сохраняем ID пользователя в состоянии
            await state.update_data(user_id=int(user_id))

            # Запрашиваем текст сообщения
            await message.answer("Введите текст сообщения:")
            await state.set_state(SendMessageState.ENTER_MESSAGE)
        else:
            # Запрашиваем текст сообщения
            await message.answer("Пользователя с таким tgId нет в базе данных:")


@router.message(SendMessageState.ENTER_MESSAGE)
async def handle_enter_message_to_one(message: Message, state: FSMContext, bot: Bot):
    """
    Обрабатывает ввод текста сообщения для отправки одному пользователю или всем.
    """

    # Получаем данные из состояния
    data = await state.get_data()
    topic = data.get("topic", "Без темы")
    user_id = data.get("user_id")
    

    # Формируем текст сообщения
    message_text = message.text

    if message.from_user.id in admins:
        status="новость" # так как это команда "Отправить сообщение от админа"
        model_name_for_messages="SendMessagesAdmin"
        fromWhomUser="Администратор"
        if user_id:
            #Отправляем пользователю, с проверкой на бан, и добавлением в базу данных в таблицу отправленные
            # user_name=message.from_user.username
            await send_message_to_user(
                                fromWhomUser=fromWhomUser,
                                bot=bot,
                                user_id=int(user_id),
                                topic=topic, 
                                message_text=message_text,
                                status=status,
                                model_name_for_messages=model_name_for_messages,
                                )
            await message.answer(f"Сообщение отправлено пользователю: {user_id}.")
        else:
            # Формируем сообщение для всех пользователей
            formatted_message = format_message_for_send(fromWhomUser, "Всем пользователям", topic, message_text)
            # Используем функцию send_messages_to_users_all для отправки
            await send_messages_to_users_all(
                bot=bot,
                message=formatted_message,
            )
            await message.answer(f"Сообщение отправлено всем пользователям.")
            #Заносим данные в таблицу отправленных сообщений от админа
            # Данные для добавления в таблицу SendMessagesAdmin
            #"tgId": 1 - потому что all_users 
            record_data = {
                "tgId": 1,
                "topic": topic,
                "body": message_text,
                "createdAt": str(get_current_time()),  # Форматируем дату и время
                "status": status,  # Статус по умолчанию
            }

            # Добавляем запись в таблицу SendMessagesAdmin
            success = await add_or_update_record(
                model_name="SendMessagesAdmin",  # Имя модели
                filters={
                    "tgId": 1, 
                    "topic": topic, 
                    "body": message_text
                    },  # Фильтры для поиска существующей записи
                data=record_data,  # Данные для добавления/обновления
            )

            if success:
                logger.info(f"Запись о сообщении добавлена/обновлена в таблице SendMessagesAdmin для всех пользователей.")
            else:
                logger.error(f"Не удалось добавить/обновить запись в таблице SendMessagesAdmin для для всех пользователей.")

        # Сбрасываем состояние
        await state.clear()
    else:
        status="не прочитано" # так как это команда "Отправить сообщение админу пользователем"
        model_name_for_messages="SendMessagesUser"
        for admin_id in admins:
            #Отправляем админу
            user_name=message.from_user.username
            fromWhomUser_tgId=int(message.from_user.id)

            await send_message_to_user(
                                fromWhomUser=user_name,
                                bot=bot,
                                user_id=int(admin_id),
                                topic=topic, 
                                message_text=message_text,
                                status=status,
                                model_name_for_messages=model_name_for_messages,
                                fromWhomUser_tgId=fromWhomUser_tgId,
                                )
            await message.answer(f"Сообщение отправлено администратору бота.")
        else:
            logger.error(f"Админа с таким Id: {admin_id} не существует.")

        # Сбрасываем состояние
        await state.clear()



@router.callback_query(F.data.startswith("mes_"))
# @router.callback_query(F.data == "incoming_mes")
async def handle_incoming_messages(callback: CallbackQuery):
    """
    Обрабатывает нажатие на кнопку "Мои входящие сообщения".
    """
    if callback.from_user.id not in admins:
        await callback.message.answer("Извините, у Вас нет прав доступа.")
        return
    
    data = callback.data.split("_")
    model_name = data[1]  # Название модели
    

    # Получаем все записи из таблицы SendMessagesAdmin
    records = await get_all_records(model_name)
    
    if not records:
        await callback.message.answer("У вас нет входящих сообщений.")
        return
    
    # Сортируем записи:
    # 1. Сначала по статусу: непрочитанные (не прочитано) сверху.
    # 2. Затем по дате: новые сверху, старые снизу.
    sorted_records = sorted(
        records,
        key=lambda x: (x.status != "не прочитано"),  # Сортировка по статусу и дате
    )

    # Формируем сообщение с количеством записей
    await callback.message.answer(f"У вас {len(records)} входящих сообщения.")


    keyboard=kb.create_list_in_messages_keyboard(sorted_records, model_name)

    # Отправляем клавиатуру
    await callback.message.answer("Выберите сообщение для просмотра деталей:", reply_markup=keyboard)

    # Подтверждаем обработку callback
    await callback.answer()


@router.callback_query(F.data.startswith("messageDetail_"))
async def handle_message_detail(callback: CallbackQuery):
    """
    Обрабатывает нажатие на кнопку с сообщением и выводит детали.
    """
    if callback.from_user.id not in admins:
        await callback.message.answer("Извините, у Вас нет прав доступа.")
        return
    
    data = callback.data.split("_")
    model_name = data[1]  # Название модели
    message_id = int(data[2])  # Извлекаем ID сообщения

    # Получаем запись из базы данных
    record = await get_record_by_id(model_name, message_id)

    if not record:
        await callback.message.answer("Сообщение не найдено.")
        return
    
    # Определяем, как назвать строку "Кому/от кого" в зависимости от статуса
    if record.status in ["ответ", "новость"]:
        recipient_info = f"Кому: {record.tgId}\n"
    elif record.status in ["прочитано", "не прочитано", "предложение"]:
        recipient_info = f"От кого: {record.tgId}\n"
    else:
        recipient_info = f"Кому/от кого: {record.tgId}\n"  # На случай, если статус неизвестен

    # Формируем текст с деталями сообщения
    detail_text = (
        f"ID: {record.id}\n"
        f"{recipient_info}\n"
        f"Тема: {record.topic}\n"
        f"Текст: {record.body}\n"
        f"Дата создания: {record.createdAt}\n"
        f"Статус: {record.status}"
    )

    if model_name=="SendMessagesUser":
        #Создаем клавиатуру с действиями для выбранного сообщения
        keyboard = kb.create_read_or_answer_keyboard(model_name, record.id)
        # Отправляем детали сообщения
        await callback.message.answer(detail_text, reply_markup=keyboard)
    else:
        await callback.message.answer(detail_text)
    
    # Подтверждаем обработку callback
    await callback.answer()


@router.callback_query(F.data.startswith("offers_"))
# @router.callback_query(F.data == "adding")
async def handle_users_suggest_adding(callback: CallbackQuery, state: FSMContext):
    """
    Обрабатывает нажатие на кнопку "Предложить добавить" у пользователя. и "Посмотреть предложения" у админа
    """
    # Получаем список моделей
    models = await get_models(combined_model_names_for_users)

    data = callback.data.split("_")
    action = data[1]  # Название действия

    if action=="suggest":
        # Получаем количество записей для всех моделей
        record_counts = await get_record_counts_for_models(combined_model_names_for_users)
    else:
        
        record_counts = await get_record_counts_for_models(analog_model_names_reserv)
        # Удаляем суффикс "Reserv" из ключей в record_counts
        record_counts = {key.replace("Reserv", ""): value for key, value in record_counts.items()}

    
    # Создаем клавиатуру с моделями
    keyboard = kb.create_models_keyboard(models, combined_model_names_for_users, action, record_counts)

    # Отправляем сообщение с клавиатурой
    await callback.message.answer("Выберите модель:", reply_markup=keyboard)

    # Подтверждаем обработку callback
    await callback.answer()

@router.callback_query(F.data.startswith("suggest_"))
async def handle_add_model(callback: CallbackQuery, state: FSMContext):
    """
    Обрабатывает выбор модели для предложения добавить - это для пользователя.
    """
    
    model_name = callback.data.replace("suggest_", "")
    await state.update_data(model_name=model_name)
    if model_name in analog_model_names:
        # Получаем список городов
        towns = await get_all_records("Town")

        # Извлекаем названия городов
        town_names = [town.town for town in towns] 

        # Создаем клавиатуру с городами
        towns_keyboard = kb.create_towns_text_keyboard(town_names)
                
        # Отправляем сообщение с клавиатурой
        await callback.message.answer(
            "Выберите город (если не видно, то скрой свою клавиатуру):",
            reply_markup=towns_keyboard
        )
        
        await state.set_state(AddAnalogModel.waiting_for_town)
    else:
        keyboard=kb.create_suggest_downloads_send_kb(model_name=model_name)
        # Отправляем сообщение с клавиатурой
        await callback.message.answer(
            "Чтобы добавить данные, нужно:\n\n"
            "1. Скачать шаблон.\n"
            "2. Заполнить по образцу.\n"
            "3. Отправить файл админу на согласование.",
            reply_markup=keyboard
)  
    await callback.answer()
    
@router.callback_query(F.data.startswith("new_"))
async def handle_add_model(callback: CallbackQuery, state: FSMContext):
    """
    Продолжает обрабатывает Посмотреть предложения - для админа".
    """
    
    if callback.from_user.id in admins:
        model_name = callback.data.replace("new_", "")
        # Получаем все записи из модели
        model_name=f'{model_name}Reserv'
        records = await get_all_records(model_name=model_name)
        if not records:
            await callback.message.answer("Предложений пока нет.")
            await callback.answer()
            return

        # Создаем клавиатуру с кнопками
        keyboard = kb.create_list_in_analog_models_keyboard(records, model_name)
        # Отправляем сообщение с клавиатурой
        await callback.message.answer("Выберите раздел:", reply_markup=keyboard)
        
        # Подтверждаем обработку callback
        await callback.answer()

    else:
        await callback.message.answer("Извините, у Вас нет прав доступа")

@router.callback_query(F.data.startswith("approve_"))
async def handle_approve(callback: CallbackQuery, state: FSMContext):
    """
    Обрабатывает нажатие на кнопку "Согласовать".
    Переносит запись из резервной таблицы в основную.
    """

    if callback.from_user.id not in admins:
        await callback.message.answer("Извините, у Вас нет прав доступа.")
        return
    
    # Извлекаем данные из callback_data
    data = callback.data.split("_")
    model_name = data[1]  # Название модели (например, "FoodReserv")
    record_id = int(data[2])  # ID записи

    # Получаем запись по ID
    record = await get_record_by_id(model_name, record_id)
    if not record:
        await callback.message.answer("Запись не найдена.")
        await callback.answer()
        return

    # Определяем имя основной таблицы (убираем "Reserv" из названия)
    main_model_name = model_name.replace("Reserv", "")

    # Переносим запись в основную таблицу
    try:
        # Копируем данные из резервной записи в основную
        record_data = {key: value for key, value in record.__dict__.items() if not key.startswith('_')  and key != 'id'}
        
        # Используем универсальную функцию для добавления или обновления записи
        check = await add_or_update_record(
            model_name=main_model_name,
            filters=record_data, 
            data=record_data
        )

        if not check:
            await callback.message.answer("Не удалось добавить запись в основную таблицу.")
            await callback.answer()
            return

        # Удаляем запись из резервной таблицы с помощью универсальной функции
        success = await remove_record(
            model_name=model_name,
            filters={"id": record_id}
        )

        if not success:
            await callback.message.answer("Не удалось удалить запись из резервной таблицы.")
            await callback.answer()
            return

        await callback.message.answer("Запись успешно перенесена в основную таблицу.")
    except Exception as e:
        logger.error(f"Ошибка при переносе записи: {e}")
        await callback.message.answer("Произошла ошибка при переносе записи.")

    # Подтверждаем обработку callback
    await callback.answer()



@router.callback_query(F.data.startswith("reject_"))
async def handle_reject(callback: CallbackQuery, state: FSMContext):
    """
    Обрабатывает нажатие на кнопку "Отменить".
    Удаляет запись из резервной таблицы и запрашивает текст объяснения.
    """
    if callback.from_user.id not in admins:
        await callback.message.answer("Извините, у Вас нет прав доступа.")
        return
    
    # Извлекаем данные из callback_data
    data = callback.data.split("_")
    model_name = data[1]  # Название модели (например, "FoodReserv")
    record_id = int(data[2])  # ID записи

    # Получаем запись по ID
    record = await get_record_by_id(model_name, record_id)
    if not record:
        await callback.message.answer("Запись не найдена.")
        await callback.answer()
        return

    # Сохраняем ID записи и tgId пользователя в состояние
    await state.update_data(
        model_name=model_name,
        record_id=record_id,
        tgId=record.tgId  # Telegram ID пользователя, который предложил запись
    )

    # Запрашиваем текст объяснения
    await callback.message.answer("Напишите причину, по которой запись была отклонена:")

    # Устанавливаем состояние ожидания текста объяснения
    await state.set_state(AddAnalogModel.waiting_for_reject_reason)

    # Подтверждаем обработку callback
    await callback.answer()


@router.message(AddAnalogModel.waiting_for_reject_reason)
async def handle_reject_reason(message: Message, state: FSMContext, bot: Bot):
    """
    Обрабатывает текст объяснения и отправляет сообщение пользователю.
    """
    if message.from_user.id not in admins:
        await message.answer("Извините, у Вас нет прав доступа.")
        return
    
    # Получаем данные из состояния
    data = await state.get_data()
    model_name = data.get("model_name")
    record_id = int(data.get("record_id"))
    tgId = data.get("tgId")  # Telegram ID пользователя

    # Удаляем запись из резервной таблицы с помощью универсальной функции
    try:
        success = await remove_record(
            model_name=model_name,
            filters={"id": record_id}  # Фильтр по ID записи
        )

        if not success:
            await message.answer("Не удалось удалить запись из резервной таблицы.")
            return
        else:
            await message.answer("Запись удалена успешно.")
        #     return

    except Exception as e:
        logger.error(f"Ошибка при удалении записи: {e}")
        await message.answer("Произошла ошибка при удалении записи.")
        return

    # Формируем сообщение для пользователя
    reason = message.text
    user_message = (
        f"Ваше предложение было отклонено.\n"
        f"Причина: {reason}"
    )
    fromWhomUser="Администратор"
    topic="Предложить добавить"
    status="предложение отклонено"
    model_name_for_messages="SendMessagesAdmin"

    # Отправляем сообщение пользователю
    try:
        await send_message_to_user(
                    fromWhomUser=fromWhomUser,   
                    bot=bot,
                    user_id=int(tgId),
                    topic=topic, 
                    message_text=user_message,
                    status=status, #новость или ответ, или предлоежние
                    # model_name=model_name, #В предложении на сообщение модель для исключения (settings) не нужна
                    model_name_for_messages=model_name_for_messages,
                )
    #     await bot.send_message(chat_id=tgId, text=user_message)
        await message.answer("Сообщение пользователю отправлено.")
    except Exception as e:
        logger.error(f"Ошибка при отправке сообщения пользователю: {e}")
        await message.answer("Не удалось отправить сообщение пользователю.")

    # Очищаем состояние
    await state.clear()

@router.callback_query(F.data.startswith("do_"))
async def handle_do_choice(callback: CallbackQuery):
    """
    Обрабатываем кнопки админа "Действия с данными, сообщения, кнопки для user. Это нужно для удобной работы с кнопками. То есть в этих кнопках будут находится другие кнопки
    """
    if callback.from_user.id not in admins:
        await callback.message.answer("Извините, у Вас нет прав доступа.")
        return
    
    # Словарь для хранения клавиатур
    keyboards = {
        "users": kb.main_users,
        "data": kb.do_admin_kb_data,
        "messages": kb.do_admin_kb_messages,
    }
    # Извлекаем название клавы из колбек
    data = callback.data.split("_")
    name_kb = data[1]  # Название модели

    # Выбираем клавиатуру из словаря
    keyboard = keyboards.get(name_kb)

    if not keyboard:
        await callback.message.answer("Клавиатура не найдена.")
        return

    # Отправляем сообщение с клавиатурой
    await callback.message.answer(
        'Выбери действие',
        reply_markup=keyboard
    )

    # Подтверждаем обработку callback
    await callback.answer()


@router.callback_query(F.data.startswith("read_"))
async def handle_read_message(callback: CallbackQuery):
    """
    Обрабатывает нажатие на кнопку "Прочитано".
    """
    if callback.from_user.id not in admins:
        await callback.message.answer("Извините, у Вас нет прав доступа.")
        return
    
    # Извлекаем данные из callback
    data = callback.data.split("_")
    model_name = data[1]  # Название модели (не используется в данном случае)
    record_id = int(data[2])  # ID записи (сообщения)

    # Обновляем статус сообщения на "прочитано"
    success = await update_message_status(record_id, "прочитано")

    if success:
        await callback.message.answer("Сообщение отмечено как прочитанное.")
    else:
        await callback.message.answer("Не удалось обновить статус сообщения.")

    # Подтверждаем обработку callback
    await callback.answer()

@router.callback_query(F.data.startswith("answer_"))
async def handle_answer_message(callback: CallbackQuery, state: FSMContext):
    """
    Обрабатывает нажатие на кнопку "Ответить". Запрашиваем у админа ввести текст письма
    """
    if callback.from_user.id not in admins:
        await callback.message.answer("Извините, у Вас нет прав доступа.")
        return

    # Извлекаем данные из callback
    data = callback.data.split("_")
    model_name = data[1]  # Название модели (не используется в данном случае)
    record_id = int(data[2])  # ID записи (сообщения)

    # Сохраняем record_id в состоянии
    await state.update_data(model_name=model_name,record_id=record_id)

    # Запрашиваем текст ответа
    await callback.message.answer("Напишите текст ответа:")
    await state.set_state(ReplyToMessage.waiting_for_reply_text)

    # Подтверждаем обработку callback
    await callback.answer()

@router.message(ReplyToMessage.waiting_for_reply_text)
async def handle_reply_text(message: Message, state: FSMContext, bot: Bot):
    """
    Обрабатывает текст ответа и отправляет его пользователю.
    """
    if message.from_user.id not in admins:
        await message.answer("Извините, у Вас нет прав доступа.")
        return
    
    # Извлекаем данные из состояния
    data = await state.get_data()
    record_id = data["record_id"]
    model_name = data["model_name"]

    # Получаем текст ответа
    reply_text = message.text

    # Получаем запись по ID
    record = await get_record_by_id(model_name, record_id)
    if not record:
        await message.answer("Запись не найдена.")
        await message.answer()
        return


    # Формируем сообщение для пользователя
    
    user_message = (
        f"Ответ на сообщение.\n"
        f"{reply_text}"
    )
    fromWhomUser="Администратор"
    topic=record.topic
    status="ответ"
    model_name_for_messages="SendMessagesAdmin"
    user_id = record.tgId  # ID пользователя, который отправил сообщение

    # Отправляем сообщение пользователю
    try:
        await send_message_to_user(
                    fromWhomUser=fromWhomUser,   
                    bot=bot,
                    user_id=user_id,
                    topic=topic, 
                    message_text=user_message,
                    status=status, #новость или ответ, или предлоежние
                    # model_name=model_name, #Ответ на сообщение модель для исключения (settings) не нужна
                    model_name_for_messages=model_name_for_messages,
                )
      
        # Обновляем статус сообщения на "прочитано"
        await update_message_status(record_id, "прочитано")

        await message.answer("Ответ отправлен.")
    except ValueError as e:
        logger.error(f"Ошибка: {e}")  

    # Обновляем статус сообщения на "прочитано"
    success = await update_message_status(record_id, "прочитано")

    if success:
        await message.answer("Сообщение отмечено как прочитанное.")
    else:
        await message.answer("Не удалось обновить статус сообщения.")      


    # Сбрасываем состояние
    await state.clear()
    

@router.callback_query(F.data == "MyOffers")
async def my_offers_main(callback: CallbackQuery):
    """
    Обработка кнопки "Mои предложения".
    Показываем 2 кнопки: Согласованные (кол-во) и На согласовании (кол-во).
    """
    tg_id = callback.from_user.id

    # Считаем количество согласованных записей (в основных моделях)
    approved_count = 0
    filters={"tgId": tg_id}
    for model_name in analog_model_names:
        # Тут ваша функция, которая возвращает кол-во записей по tgId
        cnt = await get_record_count_analog(model_name, filters)
        approved_count += cnt

    # Считаем количество не согласованных (в резервных моделях)
    pending_count = 0
    for model_name in analog_model_names_reserv:
        cnt = await get_record_count_analog(model_name, filters)
        pending_count += cnt

    # Формируем клавиатуру
    keyboard = kb.my_offers_main_kb(approved_count, pending_count)
    await callback.message.answer("Ваши предложения:", reply_markup=keyboard)
    await callback.answer()


@router.callback_query(F.data == "myoffers_approved")
async def show_approved_offers(callback: CallbackQuery):
    """
    Нажатие на "Согласованные (X)"
    Показываем список записей в моделях analog_model_names
    """
    tg_id = callback.from_user.id
    records_to_show = []

    for model_name in analog_model_names:
        # Получаем все записи пользователя из модели model_name
        # Предположим, есть метод get_records_for_user(model_name, tg_id)
        rows = await get_records_for_user(model_name, tg_id)
        if len(rows)!=0:
            for r in rows:
                # record_id = r.id
                record_id = r['id']
                record_tgId = r['tgId']
                # Как текст кнопки используем, например, r.name
                button_text = f"{analog_model_names[model_name]}: {r['name']}"
                records_to_show.append({
                    "model_name": model_name,
                    "record_id": record_id,
                    "button_text": button_text,
                    "record_tgId": record_tgId,
                })

    if records_to_show:
        keyboard = kb.my_offers_records_kb(records_to_show, is_reserv=False)
        await callback.message.answer("Согласованные записи:", reply_markup=keyboard)
    else:
        await callback.message.answer("У вас нет согласованных записей.")
    await callback.answer()

@router.callback_query(F.data == "myoffers_pending")
async def show_pending_offers(callback: CallbackQuery):
    """
    Нажатие на "На согласовании (Y)"
    Показываем список записей в моделях analog_model_names_reserv
    """
    tg_id = callback.from_user.id
    records_to_show = []

    for model_name in analog_model_names_reserv:
        rows = await get_records_for_user(model_name, tg_id)
        if len(rows)!=0:
            for r in rows:
                record_id = r['id']
                record_tgId = r['tgId']
                button_text = f"{analog_model_names_reserv[model_name]}: {r['name']}"
                records_to_show.append({
                    "model_name": model_name,
                    "record_id": record_id,
                    "button_text": button_text,
                    "record_tgId": record_tgId,
                })

    if records_to_show:
        keyboard = kb.my_offers_records_kb(records_to_show, is_reserv=True)
        await callback.message.answer("Записи на согласовании:", reply_markup=keyboard)
    else:
        await callback.message.answer("У вас нет записей на согласовании.")
    await callback.answer()

@router.callback_query(F.data.startswith("myoffers_item:"))
async def show_offer_detail(callback: CallbackQuery):
    """
    Нажатие на конкретную запись: "myoffers_item:{model_name}:{record_id}"
    Показываем детальное описание, +кнопки [Удалить] , а [Редактировать] - не стала делать
    """
    
    _, model_name, record_id_str, record_tgId_str = callback.data.split(":")
    record_id = int(record_id_str)
    record_tgId = int(record_tgId_str)

    if callback.from_user.id==record_tgId:
        # Достаем запись из БД
        record = await get_record_by_id(model_name, record_id)
        if not record:
            await callback.message.answer("Запись не найдена.")
            await callback.answer()
            return

        # Формируем текст из всех нужных полей
        message_text = get_full_info_message(record)

        # Клавиатура с &laquo;Редактировать&raquo; и &laquo;Удалить&raquo;
        keybord = kb.my_offers_detail_kb(model_name, record_id, record_tgId)
        await callback.message.answer(message_text, reply_markup=keybord)
        await callback.answer()

@router.callback_query(F.data.startswith("myoffers_delete:"))
async def delete_offer_confirm(callback: CallbackQuery):
    """
    Нажатие на &laquo;Удалить&raquo;.
    Спрашиваем подтверждение.
    Формат: "myoffers_delete:{model_name}:{record_id}"
    """
    _, model_name, record_id_str, record_tgId_str = callback.data.split(":")
    record_id = int(record_id_str)
    record_tgId = int(record_tgId_str)

    keyboard = kb.confirm_delete_kb(model_name, record_id, record_tgId)
    await callback.message.answer("Вы уверены, что хотите удалить эту запись?", reply_markup=keyboard)
    await callback.answer()

@router.callback_query(F.data.startswith("myoffers_confirm_delete:"))
async def delete_offer_final(callback: CallbackQuery):
    """
    Пользователь нажал &laquo;Да, удалить&raquo;.
    Формат: "myoffers_confirm_delete:{model_name}:{record_id}"
    Удаляем запись из БД.
    """
    _, model_name, record_id_str, record_tgId_str = callback.data.split(":")
    record_id = int(record_id_str)
    record_tgId = int(record_tgId_str)

    if callback.from_user.id==record_tgId:
        filters={
                "id": record_id,
                "tgId": record_tgId,
                }

        success = await remove_record(model_name, filters)
        if success:
            await callback.message.answer("Запись удалена.")
        else:
            await callback.message.answer("Не удалось удалить запись (возможно, её уже нет).")
        await callback.answer()

@router.callback_query(F.data.startswith("myoffers_cancel_delete"))
async def cancel_delete(callback: CallbackQuery):
    """
    Пользователь нажал &laquo;Нет, отменить&raquo; при удалении и при отмены брони, отмену действия тоже этот обработчик
    """
    await callback.message.answer("Действие отменено.")
    await callback.answer()

@router.callback_query(F.data.startswith("plus_event_"))
async def handle_parser_selection(callback: CallbackQuery):

    if callback.from_user.id not in admins:
        await callback.message.answer("Извините, у Вас нет прав доступа.")
        return
    
    # Разбираем callback_data
    data = callback.data.split("_")
    
    event_id = int(data[2])  # Ключ парсера
    
    event=await get_record_by_id("EventCheck", event_id)
    
    if event:
    #    event_dict=convert_orm_to_dict("Event", event_check)
        event_dict = {
                'town': event.town,
                'event': event.event,
                'description': event.description,
                'event_date': event.event_date,
                'time': event.time,
                'place': event.place,
                'price': event.price,
                'link_to_source': event.link_to_source,
                'action': event.action,
            }
        # Проверяем есть ли записи в это время. Сначала в фильтр добавляем дату и время и город и место
        filters_check=get_filters_for_model("EventCheck", event_dict)# Город, дата, время, место
        check_date = await get_all_in_town("Event", event.town, filters_check)
       
        if check_date:
           await callback.answer(f"Нет. Мероприятие не добавлено, в это время уже есть в данном месте мероприятие!")
        else:
            filters = get_filters_for_model("Event", event_dict)
            event_new = await add_or_update_record("Event", filters, event_dict)

            #Меняем грейд, для правильной фильтрации, чтобы потом не показывать эти мероприятия, так как они уже проверены
            await update_grade("EventCheck", event_id, 2)
            await callback.answer(f"Мероприятие '{event.event}' добавлено!")
    else:
       logger.error(f'error')   
       await callback.message.answer("Ошибка: мероприятие не найдено.") 

    # Подтверждаем обработку callback_query
    await callback.answer()

@router.callback_query(F.data.startswith("hide_event_"))
async def handle_parser_selection(callback: CallbackQuery):

    if callback.from_user.id not in admins:
        await callback.message.answer("Извините, у Вас нет прав доступа.")
        return
    
    # Разбираем callback_data
    data = callback.data.split("_")
    
    event_id = data[2]  # Ключ парсера
    
    event = await get_record_by_id("EventCheck", int(event_id))
    
    if event:
    # Меняем грейд события для того, чтобы оно больше не показывалось, так как стоит фильтр по грейду на показ мероприятий 
       await update_grade("EventCheck", int(event_id), 2)
       await callback.message.answer(f"Мероприятие '{event.event}' скрыто, здесь его больше не увидите, его можно будет посмотреть только в таблице EventCheck!")
    else:
       logger.error(f'error')   
       await callback.message.answer("Ошибка: мероприятие не найдено.") 

    # Подтверждаем обработку callback_query
    await callback.answer()

@router.callback_query(F.data.startswith("kino_"))
async def handle_kino_selection(callback: CallbackQuery):
# Обработка нажатия на кнопку кинотеатры

    # Разбираем callback_data
    data = callback.data.split("_")
    
    town = data[1]  # Ключ парсера
    if town =="Грязи":
        records= await get_all_records("ImageModel")

        if records:
            for img in records:
                await callback.answer("Афиша с киносеансами подготавливается, ждите.")
                # Создаем временный файл для изображения
                temp_file_name=f"temp_{img.file_name}"
                file_data=img.file_data
                with open(temp_file_name, "wb") as temp_file:
                    temp_file.write(file_data)
                
                # Отправляем изображение пользователю
                photo = FSInputFile(temp_file_name)
                await callback.message.answer_photo(photo)
                
                # Удаляем временный файл
                os.remove(temp_file_name)

                await callback.message.answer(f"Источник: {img.site}")
        else:
            await callback.message.answer("В данный момент нет информации по Афише с киносеансами, попробуйте позже.")
    else:
        await callback.message.answer("Этот город не поддерживается.")

# @router.callback_query(F.data.startswith("filter_date_"))
# async def handle_filter_date(callback: CallbackQuery):
#     """
#     Обрабатывает выбор фильтрации по дате.
#     """
#     town = callback.data.split("_")[2]
#     # Здесь можно вызвать календарь для выбора даты
#     await callback.message.answer("Выберите дату (реализация календаря зависит от библиотеки).")

@router.callback_query(F.data.startswith("filter_place_"))
async def handle_filter_place(callback: CallbackQuery):
    """
    Обрабатывает выбор фильтрации по месту проведения.
    """
    town = callback.data.split("_")[2]
    date_day_today = get_today_date_dmy()

    # Получаем мероприятия
    events = await get_events_future_by_town(town=town, event_date=date_day_today)


    if events:
        # Группируем мероприятия по месту проведения
        places = get_places_from_events(events)

        # Создаем клавиатуру с местами проведения
        keyboard = kb.create_events_one_place_keyboard(places=places, town=town)

        await callback.message.answer("Выберите место проведения:", reply_markup=keyboard.as_markup())
    else:
        await callback.message.answer("Мероприятий не найдено.")


@router.callback_query(F.data.startswith("place_"))
async def handle_place_selection(callback: CallbackQuery):
    """
    Обрабатывает выбор места проведения.
    """
    # Извлекаем id мероприятий из callback_data
    event_id = int(callback.data.split("_")[1])
    town = callback.data.split("_")[2]

    event_first = await get_event_by_id(event_id)

    place=event_first.place

    date_day_today = get_today_date_dmy()

    # Получаем мероприятия
    events = await get_events_future_by_town(town=town, event_date=date_day_today, place=place)
    
    
    if events:
        # Сортируем мероприятия по дате
        sorted_events = sorted(
            events,
            key=lambda x: (x['event_date'])  # Используем ключи словаря
        )

        message = format_data_message(sorted_events, "Мероприятия по выбранному месту")
        if message =="Сообщение слишком длинное. Хотите скачать Excel-файл со всеми записями?":
            for event in sorted_events:
                message_text = format_data_one_message(event)
                # Отправляем сообщение
                await callback.message.answer(message_text)
        else:
            message_text = message
            await callback.message.answer(message_text)
    else:
        await callback.message.answer("Мероприятий не найдено.")

# Обработка нажатия кнопки добавить шаблон
@router.callback_query(F.data == "addTemplate")
async def handle_parse_section(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in admins:
        await callback.message.answer("Извините, у Вас нет прав доступа.")
        return
    
    # Отправляем клавиатуру с городами
    await callback.message.answer("Пожалуйста, загрузите файл шаблона в формате .xlsx.")
    # Устанавливаем состояние ожидания файла
    await state.set_state(Form.waiting_for_template)  


# Обработчик загрузки файла
@router.message(F.content_type == ContentType.DOCUMENT, Form.waiting_for_template)
async def process_file(message: Message, state: FSMContext):
    if message.from_user.id not in admins:
        await message.answer("Извините, у Вас нет прав доступа.")
        return

    # Проверяем, что файл имеет расширение .xlsx
    if message.document.mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        file_id = message.document.file_id
        file = await message.bot.get_file(file_id)
        file_path = file.file_path

        # Скачиваем файл
        downloaded_file = await message.bot.download_file(file_path)

        # Читаем файл в бинарном формате
        file_data = downloaded_file.read()

        file_name=message.document.file_name
        file_data=file_data
        # createdAt=datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Добавляем или обновляем запись
        success = await add_or_update_file_record(
            model_name="TemplateModel",
            file_name=file_name,
            file_data=file_data,
            filters={"file_name": file_name},  # Опционально: ищем запись по имени файла
        )

        if success:
            await message.answer("Файл успешно загружен и сохранен в базу данных.")
        else:
            await message.answer("Произошла ошибка при загрузке шаблона в таблицу.")
    else:
        await message.answer("Пожалуйста, загрузите файл в формате .xlsx.")

    # Сбрасываем состояние
    await state.clear()

@router.callback_query(F.data.startswith("template_"))
async def handle_download_template(callback: CallbackQuery):
    # Извлекаем название модели из callback_data
    model_name = callback.data.split("_")[1]

    template = await get_template(model_name=model_name)
    try:
        if not template:
            await callback.message.answer("Шаблон не найден.", show_alert=True)
            return

        # Преобразуем бинарные данные в файл
        file_data = BytesIO(template.file_data)
        file_data.name = template.file_name  # Указываем имя файла

        # Отправляем файл пользователю
        await callback.message.answer_document(
            document=BufferedInputFile(file_data.getvalue(), filename=file_data.name),
            # document=FSInputFile(file_data),
            caption=f"Вот ваш шаблон."
        )

        # Закрываем BytesIO
        file_data.close()
    except Exception as e:
        logger.error(f"Ошибка при получении файла-шаблона: {e}")
        await callback.message.answer("Не удалось получить файл-шаблон, сообщите от ошибке админу.")


# Обработка нажатия кнопки добавить шаблон
@router.callback_query(F.data == "sendForApproval")
async def handle_parse_section(callback: CallbackQuery, state: FSMContext):
    user_name=callback.from_user.full_name
    user_id=callback.from_user.id
    # Сохраняем user_name в состояние
    await state.update_data(user_name=user_name)
    # Сохраняем user_id в состояние
    await state.update_data(user_id=user_id)
    
    # Устанавливаем состояние ожидания файла
    await state.set_state(Form.waiting_for_ready)
    
    await callback.message.answer("Пожалуйста, загрузите файл шаблона в формате .xlsx.") 

# Обработчик загрузки файла
@router.message(F.content_type == ContentType.DOCUMENT, Form.waiting_for_ready)
async def process_file(message: Message, state: FSMContext, bot: Bot):
    # if message.from_user.id not in admins:
    #     await message.answer("Извините, у Вас нет прав доступа.")
    #     return

    # Получаем данные из состояния
    data = await state.get_data()
    user_name = data.get("user_name")  # Извлекаем user_name
    user_id = data.get("user_id")  # Извлекаем user_id

    # Проверяем, что файл имеет расширение .xlsx
    if message.document.mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        file_id = message.document.file_id
        file = await message.bot.get_file(file_id)
        file_path = file.file_path

        # Скачиваем файл
        downloaded_file = await message.bot.download_file(file_path)

        # Читаем файл в бинарном формате
        file_data = downloaded_file.read()

        file_name=message.document.file_name
        file_data=file_data
        # createdAt=datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Преобразуем бинарные данные в файл
        file_data = BytesIO(file_data)
        file_data.name = file_name  # Указываем имя файла

        for admin_id in admins:
            try:
                await bot.send_document(
                    admin_id,
                    document=BufferedInputFile(file_data.getvalue(), filename=file_data.name),
                    # document=FSInputFile(file_data),
                    caption=f"Пришло предложение от пользователя: {user_name} его tgId: {user_id}"
                    
                )
            except TelegramForbiddenError as tg_error:
                logger.error(f"Не удалось отправить сообщение админу {admin_id}: {tg_error}")

        

        # Закрываем BytesIO
        file_data.close()

        
    else:
        await message.answer("Пожалуйста, загрузите файл в формате .xlsx.")

    # Сбрасываем состояние
    await state.clear()


@router.callback_query(F.data.startswith("filter_date_"))
async def handle_filter_place(callback: CallbackQuery):
    """
    Обрабатывает выбор фильтрации по дате.
    """
    town = callback.data.split("_")[2]
    date_day_today = get_today_date_dmy()

    # Получаем мероприятия
    events = await get_events_future_by_town(town=town, event_date=date_day_today)


    if not events:
        # Если мероприятий нет, отправляем сообщение
        await callback.message.answer(f"В городе {town} нет запланированных мероприятий.")
    else:
        
        # Сортируем мероприятия по дате
        sorted_events = sorted(
            events,
            key=lambda x: (x['event_date'])  # Используем ключи словаря
        )

        # Группируем мероприятия по дате
        events_by_date = {}
        for event in sorted_events:
            event_date = event["event_date"].strftime("%d.%m.%Y")
            if event_date not in events_by_date:
                events_by_date[event_date] = []
            events_by_date[event_date].append(event)

        # Создаем клавиатуру с датами мероприятий
        keyboard = kb.create_events_one_days_keyboard(events_by_date=events_by_date, town=town)

        
        # Отправляем клавиатуру с датами
        await callback.message.answer(
            f"Выберите дату для просмотра мероприятий в городе {town}:",
            reply_markup=keyboard.as_markup()
        )

    # Закрываем всплывающее уведомление
    await callback.answer()

@router.callback_query(F.data.startswith("show_events_"))
async def on_show_events_button_click(callback: CallbackQuery):
    """
    Обработчик нажатия на кнопку с датой.
    """
    # Извлекаем город и дату из callback_data
    data = callback.data.split("_")
    town = data[2]
    event_date_str = data[3]

    # Преобразуем строку даты в объект date
    event_date = datetime.strptime(event_date_str, "%d.%m.%Y").date()

    # Получаем мероприятия на выбранную дату
    events = await get_events_by_date_and_town(town=town, event_date=event_date)

    if not events:
        await callback.message.answer(f"На {event_date_str} в городе {town} нет мероприятий.")
    else:
        message = format_data_message(events, f"Мероприятия на {event_date_str} в городе {town}")
        if message =="Сообщение слишком длинное. Хотите скачать Excel-файл со всеми записями?":
            for event in events:
                message_text = format_data_one_message(event)
                # Отправляем сообщение
                await callback.message.answer(message_text)
        else:
            message_text = message
            await callback.message.answer(message_text)
    
        # # Формируем текст с мероприятиями
        # text = f"Мероприятия в городе {town} на {event_date_str}:\n\n"
        # for event in events:
        #     text += (
        #         f"📅 {event['event']}\n"
        #         f"⏰ {event['time']}\n"
        #         f"📍 {event['place']}\n"
        #         f"💵 {event['price']}\n"
        #         f"🔗 {event['link_to_source']}\n\n"
        #     )

        # # Отправляем сообщение с мероприятиями
        # await callback.message.answer(text)

    # Закрываем всплывающее уведомление
    await callback.answer()


@router.callback_query(F.data.startswith("my_room:"))
async def offer_my_room(callback: CallbackQuery):
    """
    Нажатие на Личный кабинет;.
    Рассказываем куда он попал и что может сделать.
    Формат: "my_room:{model_name}:{record_id}:{record_tgId}"
    """
    _, model_name, record_id_str, record_tgId_str = callback.data.split(":")
    record_id = int(record_id_str)
    record_tgId = int(record_tgId_str)

    keyboard = kb.create_suggest_downloads_send_kb(model_name, record_id, record_tgId)
    await callback.message.answer(f"Здесь Вы можете загрузить свое расписание, чтобы к Вам могли записаться или ознакомиться с Вашим расписанием в любое время дня и ночи, не отрывая Вас от важных дел. Ваше расписание будет доступно каждому по ссылке: {url_recording}/{model_name}/{record_id}", disable_web_page_preview=True, reply_markup=keyboard)
    await callback.answer()

# Обработка нажатия кнопки добавить шаблон
@router.callback_query(F.data.startswith("pushRecording_"))
async def handle_push_recording(callback_query: CallbackQuery, state: FSMContext):
    # Извлекаем данные из callback_data
    data = callback_query.data.split("_")
    model_name = data[1]
    record_id = int(data[2])
    record_tgId = int(data[3])
    if callback_query.from_user.id == record_tgId:
        await state.update_data(model_name=model_name)
        await state.update_data(record_id=record_id)
        await state.update_data(record_tgId=record_tgId)

        # Запрашиваем файл у пользователя
        await callback_query.message.answer("Пожалуйста, загрузите файл в формате xlsx.")
        await state.set_state(TemplateRecord.waiting_for_file)
        await callback_query.answer()
    else:
        # Нет прав 
        await callback_query.message.answer("Извините, у Вас нет прав доступа.")



@router.message(TemplateRecord.waiting_for_file, F.document)
async def handle_upload_file(message: Message, state: FSMContext, bot: Bot):
    """
    Обрабатывает загрузку Excel-файла c расписанием для брони и добавляет данные в базу данных.
    """
    # Получаем данные из состояния
    data = await state.get_data()
    model_name = data["model_name"]
    record_id = data["record_id"]
    record_tgId = data["record_tgId"]

    # Скачиваем файл
    file_id = message.document.file_id
    file = await bot.get_file(file_id)
    file_path = file.file_path

    # Скачиваем файл в память
    downloaded_file = await bot.download_file(file_path)
    file_bytes = downloaded_file.read()

    # Счетчик добавленных записей
    added_count = 0

    try:
        # Читаем Excel-файл с помощью openpyxl
        workbook = load_workbook(filename=BytesIO(file_bytes))
        sheet = workbook.active  # Получаем активный лист

        # Получаем заголовки столбцов (первая строка)
        headers = [cell.value for cell in sheet[1]]

        logger.info(f'headers: {headers}')

        model = "Record"

        # Получаем колонки модели
        model_columns = get_model_columns(model)

        logger.info(f'model_columns: {model_columns}')

        
        # Обрабатываем строки Excel-файла
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовки

            # Пропускаем пустые строки
            if all(cell is None or cell == "" for cell in row):
                continue  # Пропускаем пустую строку


            row_data = dict(zip(headers, row))  # Создаем словарь из заголовков и значений

            

            # Преобразуем дату и время в строки, если это необходимо
            if "time_slot_start" in row_data and isinstance(row_data["time_slot_start"], time):
                row_data["time_slot_start"] = row_data["time_slot_start"].strftime("%H:%M")  # Формат: HH:MM
            if "time_slot_finish" in row_data and isinstance(row_data["time_slot_finish"], time):
                row_data["time_slot_finish"] = row_data["time_slot_finish"].strftime("%H:%M")  # Формат: HH:MM

            # Преобразуем param в строку, если это необходимо
            for param in ["params_1", "params_2", "params_3", "params_4", "params_5"]:
                if param in row_data and isinstance(row_data[param], (int, float)):
                    row_data[param] = str(row_data[param])  # Преобразуем в строку

            # Преобразуем param в строку, если это необходимо
            for param in ["question_1", "question_2", "question_3", "question_4", "question_5"]:
                if param in row_data and isinstance(row_data[param], (int, float)):
                    row_data[param] = str(row_data[param])  # Преобразуем в строку

                

            # Добавляем дополнительные данные
            row_data["model"] = model_name
            row_data["tgId"] = int(record_tgId)
            row_data["offerId"] = int(record_id)
            url_record_my=f'{url_recording}/{model_name}/{record_id}'
            row_data["url_record_my"] = url_record_my

            # Фильтруем row_data, оставляя только те ключи, которые есть в параметрах функции
            filtered_data = {k: v for k, v in row_data.items() if k in model_columns}

            # Устанавливаем number_of_seats = 1 перед добавлением записи

            filtered_data["number_of_seats"] = 1

            # Обрабатываем колонку number_of_seats
            number_of_seats = row_data.get("number_of_seats", 1)
            
            if number_of_seats > 1:
                # Копируем строку (number_of_seats - 1) раз
                for _ in range(number_of_seats - 1):
                    
                    success = await add_record(
                        model_name = "Record",
                        data = filtered_data
                    )
                    if success:
                        added_count += 1

            
            # Добавляем оригинальную строку
            success = await add_record(
                model_name = "Record",
                data=filtered_data
            )
            if success:
                added_count += 1  # Увеличиваем счетчик добавленных записей
                logger.info(f'Данные успешно добавлены в базу')

        # Формируем сообщение о количестве добавленных записей
        logger.info(f'Формируем сообщение о добавленных записях')
        if added_count > 0:
            await message.answer(f"Данные успешно загружены! Добавлено {added_count} записей. Ваше расписание доступно по ссылке: {url_record_my}", disable_web_page_preview=True)

    except Exception as e:
        logger.error(f"Ошибка при загрузке данных: {e}")
        await message.answer(f"Произошла ошибка при загрузке данных: Проверьте, чтобы все лишние данные были удалены из таблицы (все красные комментарии и т.п.). Попробуйте снова или свяжитесь с админом.")

    # Очищаем состояние
    await state.clear()


# Обработка нажатия кнопки Посмотреть расписание
@router.callback_query(F.data.startswith("viewMyRecording_"))
async def handle_push_recording(callback_query: CallbackQuery, state: FSMContext):
    # Извлекаем данные из callback_data
    data = callback_query.data.split("_")
    model_name = data[1]
    record_id = int(data[2])
    record_tgId = int(data[3])
    if callback_query.from_user.id == record_tgId:
        # await state.update_data(model_name=model_name)
        # await state.update_data(record_id=record_id)
        # await state.update_data(record_tgId=record_tgId)

        keyboard=kb.create_action_get_record_kb(model_name=model_name, record_id=record_id, record_tgId=record_tgId)

        # Запрашиваем файл у пользователя
        await callback_query.message.answer("Выбери:", reply_markup=keyboard)
        
        await callback_query.answer()
    else:
        # Нет прав 
        await callback_query.message.answer("Извините, у Вас нет прав доступа.")
    
# # Обработка нажатия кнопки Посмотреть здесь
# @router.callback_query(F.data.startswith("getRecords_"))
# async def handle_push_recording(callback_query: CallbackQuery, state: FSMContext):
#     # Извлекаем данные из callback_data
#     data = callback_query.data.split("_")
#     model_name = data[1]
#     record_id = int(data[2])
#     record_tgId = int(data[3])
#     if callback_query.from_user.id == record_tgId:
#         # await state.update_data(model_name=model_name)
#         # await state.update_data(record_id=record_id)
#         # await state.update_data(record_tgId=record_tgId)

#         model="Record"
#         filters ={
#             "tgId": callback_query.from_user.id,
#             "model": model_name,
#             "offerId": record_id,
#         }

#         records = await get_all_records(model_name=model, filters=filters)
#         logger.info(f'records: {records}')

#         fields_to_include={
#             "id", "date_booking", "time_slot_start", "time_slot_finis", "number_of_seats"
#         }

#         model = get_model_by_name(model)
#         message_text = await generate_message_from_model(model=model, records=records, fields_to_include=fields_to_include)
#         if message_text == "Сообщение слишком длинное. Хотите скачать Excel-файл со всеми записями?":
#             for record in records:
#                 message_text = await generate_message_from_model(model=model, record=record, fields_to_include=fields_to_include)
#                 await callback_query.message.answer(message_text, parse_mode="HTML")
#         else:
#             await callback_query.message.answer(message_text, parse_mode="HTML")

#         await callback_query.answer()
#     else:
#         # Нет прав 
#         await callback_query.message.answer("Извините, у Вас нет прав доступа.")


@router.callback_query(F.data.startswith("downloadRecord_"))
async def downloadRecord(callback: CallbackQuery):
    """
    Формирует и отправляет Excel-файл с данными о записях бронях Загрузить фалй.
    [InlineKeyboardButton(text="Скачать файлом", callback_data=f"downloadRecord_{model_name}_{record_id}_{record_tgId}")],
    """
    data = callback.data.split("_")
    model_name = data[1]
    record_id = int(data[2])
    record_tgId = int(data[3])

    if callback.from_user.id == record_tgId:
        await callback.message.answer(
            "Файл подготавливается, ожидайте")
        
        model="Record"
        filters ={
            "tgId": callback.from_user.id,
            "model": model_name,
            "offerId": record_id,
        }

        records = await get_all_records(model_name=model, filters=filters)
        
        
        if records:
            # Подготавливаем данные для Excel с использованием prepare_model_data
            data = []
            headers = None

            # Сортируем по дате
            sorted_records = sorted(
                records,
                key=lambda x: (x.date_booking)  # Используем ключи словаря
            )

            if sorted_records:
                records = sorted_records
                logger.info("Сортировка записей\броней прошла усешно")
            else: 
                records = records

            for record in records:
                # Преобразуем запись в словарь (если это ORM-объект)
                raw_data = record.__dict__ if hasattr(record, "__dict__") else record

                model="Record"
                # Подготавливаем данные для модели
                prepared_data = prepare_model_data(model_name=model, raw_data=raw_data)

                # Формируем заголовки, если они еще не заданы
                if headers is None:
                    headers = list(prepared_data.keys())

                # Добавляем строку данных
                data.append(list(prepared_data.values()))

            # Формируем имя файла
            filename = f"{model_name.lower()}.xlsx"

            # Отправляем Excel-файл
            await create_and_send_excel(
                callback,
                data,
                headers,
                filename,
                model_name,  # Название листа (например, "User", "Event", "BusSchedule")
                f"Данные"  # Заголовок
            )
        else:
            await callback.message.answer(f"Данные для модели {model_name} не найдены.")
            await callback.answer()
      
    else:
        # Нет прав 
        await callback.message.answer("Извините, у Вас нет прав доступа.")

@router.callback_query(F.data.startswith("removeRecord_"))
async def removeRecord(callback: CallbackQuery):
    """
    Кнопка удалить расписание. Возвращает клавиатуру с вопросом об уверенности своих дейтсвий

    [InlineKeyboardButton(text="Удалить всё расписание", callback_data=f"removeRecord_{model_name}_{record_id}_{record_tgId}")]
    """
    data = callback.data.split("_")
    model_name = data[1]
    record_id = int(data[2])
    record_tgId = int(data[3])

    if callback.from_user.id == record_tgId:
        
        keyboard=kb.create_remove_record_kb(model_name=model_name, record_id=record_id, record_tgId=record_tgId)

        # Запрашиваем файл у пользователя
        await callback.message.answer(f"Вы уверены, что хотите удалить всё свое расписание для модели {model_name}. Скачайте себе расписание для дальнейшего форматирования и загрузки уже нового файла-расписания:", reply_markup=keyboard)
        
        await callback.answer()
      
    else:
        # Нет прав 
        await callback.message.answer("Извините, у Вас нет прав доступа.")

@router.callback_query(F.data.startswith("removeYes_"))
async def removeYes_(callback: CallbackQuery):
    """
    Кнопка удалить расписание. Возвращает клавиатуру с вопросом об уверенности своих дейтсвий

    [InlineKeyboardButton(text="Удалить, да!", callback_data=f"removeYes_{model_name}_{record_id}_{record_tgId}")],
    """
    data = callback.data.split("_")
    model_name = data[1]
    record_id = int(data[2])
    record_tgId = int(data[3])

    if callback.from_user.id == record_tgId:
        
        model="Record"
        filters ={
            "tgId": callback.from_user.id,
            "model": model_name,
            "offerId": record_id,
        }

        records_remove = await remove_records_all(model_name=model, filters=filters)
        
        if records_remove:
            await callback.message.answer(f"Данные успешно удалены.")
            await callback.answer()
        else:
            await callback.message.answer(f"Данные не найдены. Ошибка удаления. Обратитесь к админу")
            await callback.answer()
      
    else:
        # Нет прав 
        await callback.message.answer("Извините, у Вас нет прав доступа.")

@router.callback_query(F.data.startswith("filtersDataBookings_"))
async def filtersDataBookings(callback: CallbackQuery):
    """
    Возвращает даты, на которые есть бронь
    """
    data = callback.data.split("_")
    model_name = data[1]
    record_id = int(data[2])
    record_tgId = int(data[3])
    

    if callback.from_user.id == record_tgId:
        

        model="Record"
        filters ={
            "tgId": callback.from_user.id,
            "model": model_name,
            "offerId": record_id,
        }

        records = await get_all_records(model_name=model, filters=filters)
        
        if not records:
            await callback.message.answer(f"К Вам никто не записан")
        

        else:
            ## Сортируем по дате
            sorted_records = sorted(
                records,
                key=lambda x: (x.date_booking)  # Используем ключи словаря
            )

            if sorted_records:
                records = sorted_records
                logger.info("Сортировка записей\броней прошла усешно")
            else: 
                records = records

            # Получаем сегодняшнюю дату
            date_day_today = datetime.now().date()
            logger.info(f'date_day_today: {date_day_today}')

            # Группируем по дате
            records_by_date = {}
            for event in sorted_records:
                # Проверяем, что userName не пустой
                if not event.userName or event.userName.strip() == "":
                    # logger.info(f"Пропущена запись с пустым userName: {event}")
                    continue  # Пропускаем запись с пустым userName
                
                
                # Проверяем, что дата записи больше или равна сегодняшней
                if event.date_booking >= date_day_today:
                    date_booking = event.date_booking.strftime("%d.%m.%Y")
                    logger.info(f'date_booking: {date_booking}')

                    if date_booking not in records_by_date:
                        records_by_date[date_booking] = []
                    records_by_date[date_booking].append(event)

            
            # Если есть записи на будущие даты
            if records_by_date:
                

                # Создаем клавиатуру с датами мероприятий
                keyboard = kb.create_events_one_days_keyboard(events_by_date=records_by_date, model_name = model_name, record_id = record_id, record_tgId = record_tgId)

                
                # Отправляем клавиатуру с датами
                await callback.message.answer(
                    f"У Вас есть записи на такие даты. Выберите дату:",
                    reply_markup=keyboard.as_markup()
                )
            else:
                # Если нет записей на будущие даты
                await callback.message.answer("У Вас нет записей на будущие даты.")
    else:
        # Нет прав
        await callback.message.answer("Извините, у Вас нет прав доступа.")        

    # Закрываем всплывающее уведомление
    await callback.answer()

@router.callback_query(F.data.startswith("viewAllBookings_"))
async def viewAllBookings(callback: CallbackQuery):
    """
    Кнопка Брони/записи. Возвращает клавиатуру с вопросом об уверенности своих дейтсвий

    [InlineKeyboardButton(text="Удалить, да!", callback_data=f"removeYes_{model_name}_{record_id}_{record_tgId}")],
    """
    data = callback.data.split("_")
    model_name = data[1]
    record_id = int(data[2])
    record_tgId = int(data[3])
    date_booking_str = data[4]  # Получаем дату как строку

    # Преобразуем строку в объект date
    try:
        date_booking = datetime.strptime(date_booking_str, "%d.%m.%Y").date()
    except ValueError:
        await callback.message.answer("Некорректный формат даты.")
        return
    
    

    if callback.from_user.id == record_tgId:
        
        model="Record"
        filters ={
            "tgId": int(callback.from_user.id),
            "date_booking": date_booking,
            "model": model_name,
            "offerId": int(record_id),
        }

        records = await get_all_records(model_name=model, filters=filters)
        
        if records:
            #Считаем сколько броней всего
            
            count_booking = 0
            for record in records:
                if record.userName:  # Проверяем, что поле userName не пустое
                    count_booking += 1
            # Создаем клавиатуру с отфильтрованными записями
            keyboard = await kb.create_booking_keyboard(records)
            await callback.message.answer(f"{date_booking} у Вас {count_booking} записей/броней:", reply_markup=keyboard)
        else:
            await callback.message.answer("Записей не найдено.")
        
        await callback.answer()
    else:
        # Нет прав
        await callback.message.answer("Извините, у Вас нет прав доступа.")


@router.callback_query(F.data.startswith("bookingInfo_"))
async def bookingInfo(callback: CallbackQuery):
    """
    Обработчик для отображения полной информации о записи и кнопок действий.
    """
    data = callback.data.split("_")
   
    record_id = int(data[1])
    record_tgId = int(data[2])
    

    if callback.from_user.id == record_tgId:
        model_name="Record"

        # Получаем запись по ID
        record = await get_record_by_id(model_name=model_name, record_id=record_id)
        if not record:
            await callback.message.answer("Запись не найдена.")
            return

       
        message_text = get_message_bookig(data=record)

        user = await get_numberPhone(record.numberPhone)

        if user:
            # Если номер найден в таблице User
            message_text += "\n\n✔️ Пользователь с таким номером есть в боте, поэтому ему придут напоминание о броне. А Вам придет уведомление о подтверждении или отмене брони."
        else:
            # Если номер не найден в таблице User
            message_text += "\n\n❗ Внимание! Этот номер не закреплен. Рекомендуем проверить пользователя этого номера. Он не получит напоминание о брони."

        # Создаем клавиатуру с кнопками действий
        keyboard = kb.create_records_keyboard(model_name="Record", record_id=record.id, record_tgId=int(callback.from_user.id))

        await callback.message.answer(message_text, reply_markup=keyboard, parse_mode="Markdown")
        await callback.answer()
    else:
        # Нет прав
        await callback.message.answer("Извините, у Вас нет прав доступа.")


@router.callback_query(F.data.startswith("cancelRecord_"))
async def cancel_record(callback: CallbackQuery):
    """
    Обрабатывает кнопку "Отменить запись/бронь".
    Сначала запрашивает подтверждение, затем очищает поля userName и numberPhone.
    """
    data = callback.data.split("_")
    if len(data)>2:
        model_name = data[1]
        record_id = int(data[2])
        record_tgId = int(data[3])

        if callback.from_user.id == record_tgId:
            # Получаем данные о записи
            model = "Record"
            # Получаем запись по ID
            record = await get_record_by_id(model_name=model, record_id=record_id)
            if not record:
                await callback.message.answer("Запись не найдена.")
                return

            if record:
                
                # Формируем сообщение с подтверждением
                confirmation_message = (
                    f"Вы уверены, что хотите отменить запись для клиента {record.userName} "
                    f"на {record.date_booking} в {record.time_slot_start}?"
                )

                # Создаем клавиатуру с подтверждением
                keyboard = kb.create_yes_or_no_cancel_booking_keyboard(model_name=model_name, record_id=record_id, record_tgId=record_tgId)

                await callback.message.answer(confirmation_message, reply_markup=keyboard)
            else:
                await callback.message.answer("Запись не найдена.")
        else:
            await callback.message.answer("Извините, у Вас нет прав доступа.")
    else:
        tg_id=callback.from_user.id
        user=await get_user_by_tg_id(tg_id)
        model = "Record"
        record_id=int(data[1])
        # Получаем запись по ID
        record = await get_record_by_id(model_name=model, record_id=record_id)
        if not record:
                await callback.message.answer("Запись не найдена.")
                return
        
        if user.numberphone==record.numberPhone:
            
                # Формируем сообщение с подтверждением
                confirmation_message = (
                    f"Вы уверены, что хотите отменить запись"
                    f"на {record.date_booking} в {record.time_slot_start}?"
                )

                # Создаем клавиатуру с подтверждением
                keyboard = kb.user_yes_or_no_cancel_booking_keyboard(record_id=record_id)

                await callback.message.answer(confirmation_message, reply_markup=keyboard)
        else:
            await callback.message.answer("Извините, у Вас нет прав доступа.")

    
    await callback.answer()

@router.callback_query(F.data.startswith("confirmCancel_"))
async def confirm_cancel(callback: CallbackQuery):
    """
    Обрабатывает подтверждение отмены записи.
    Очищает поля userName и numberPhone.
    """
    data = callback.data.split("_")
    if len(data)>2:
        model_name = data[1]
        record_id = int(data[2])
        record_tgId = int(data[3])


        if callback.from_user.id == record_tgId:

            record = await get_record_by_id(model_name="Record", record_id=record_id)
           
                
            if record.numberPhone:
                filters={
                    "numberphone": record.numberPhone
                }
                user = await get_all_records(model_name="User", filters=filters)
                
                if user:
                    # Отправляем уведомление пользователю с tgId
                    message_text = "❌ Запись отменена!\n\n"
                    message_text += get_message_bookig(record)
                        
                    try:
                        await callback.bot.send_message(
                            chat_id=user[0].tg_id,
                            text=message_text
                        )
                    except Exception as e:
                            logger.error(f"Ошибка при отправке уведомления: {e}")

            success = await clear_booking_fields(model_name=model_name, record_id=record_id)

            if success:
                await callback.message.answer("Запись успешно отменена.")
                

            else:
                await callback.message.answer("Произошла ошибка при отмене записи.")
        else:
            await callback.message.answer("Извините, у Вас нет прав доступа.")
    else:
        tg_id=callback.from_user.id
        user=await get_user_by_tg_id(tg_id)
        model = "Record"
        record_id=int(data[1])
        # Получаем запись по ID
        record = await get_record_by_id(model_name=model, record_id=record_id)
        if not record:
                await callback.message.answer("Запись не найдена.")
                return
        
        if user.numberphone==record.numberPhone:
            success = await clear_booking_fields(model_name=model, record_id=record_id)

            if success:
                await callback.message.answer("Запись успешно отменена.")
                # Отправляем уведомление хозяину записи
                message_text = "❌ Запись отменена пользователем!\n\n"
                message_text += get_message_bookig(record)
                 
                try:
                    await callback.bot.send_message(
                        chat_id=record.tgId,
                        text=message_text
                    )
                except Exception as e:
                    logger.error(f"Ошибка при отправке уведомления: {e}")

            else:
                await callback.message.answer("Произошла ошибка при отмене записи.")
        else:
            await callback.message.answer("Извините, у Вас нет прав доступа.")

    await callback.answer()

@router.callback_query(F.data.startswith("addRecord_"))
async def addRecord_(callback: CallbackQuery):
    """
    Обрабатывает кнопку "Добавить такой же слот(окошко)".
    Копирует запись, но оставляет поля userName и numberPhone пустыми.
    """
    data = callback.data.split("_")
    model_name = data[1]
    record_id = int(data[2])
    record_tgId = int(data[3])

    if callback.from_user.id == record_tgId:
        model="Record"
        # Копируем запись
        success = await copy_record_with_empty_user_data(model_name=model, record_id=record_id)

        if success:
            await callback.message.answer("Слот успешно добавлен.")
        else:
            await callback.message.answer("Произошла ошибка при добавлении слота.")
    else:
        await callback.message.answer("Извините, у Вас нет прав доступа.")

    await callback.answer()

@router.callback_query(F.data.startswith("instruction_"))
async def instruction(callback: CallbackQuery):
    """
    Обрабатывает кнопку "Инструкция" в личном кабинете.
    
    """
    data = callback.data.split("_")
    url = f'{data[1]}_{data[2]}'

    text = (
    "📋 <b>Инструкция добавления расписания:</b>\n\n"
    "1. <b>Скачать шаблон.</b>\n"
    "2. <b>Заполнить шаблон</b>, следуя рекомендациям, которые в нем зафиксированы. Там же в файле-шаблоне можно посмотреть примеры заполнения под разные варианты (бьюти-сфера, секции, МК, репетитор, бронирование столиков в кафе и ресторане).\n\n"
    "   - <b>Лист 1 - Пояснения по столбцам</b> (инструкция по столбцам, тоже самое, что и инструкция на Листе 2).\n"
    "   - <b>Лист 2 - Инструкция</b>.\n"
    "   - <b>Лист 3, 4, 5, 6, 7 - Примеры заполнения</b>.\n"
    "3. <b>Отправить готовый, заполненный шаблон</b>, нажав на кнопку \"Отправить\".\n"
    f"4. <b>Всё! Готово!</b> Теперь Ваше расписание увидят люди по ссылке: {url}\n\n"
    "📅 <b>Клиенты бронируют свободный слот</b> (окошко, стол и пр.), а Вам приходит уведомление, что к Вам записались на такую-то дату, на такое-то время.\n\n"
    "📋 <b>Инструкция работы с личным кабинетом:</b>\n\n"
    "1. <b>Пройдя по кнопке \"Расписание\"</b>, Вы сможете:\n"
    "   - <b>Скачать</b> свое расписание, как оно выглядит на данный момент (с занятыми и свободными слотами).\n"
    "   - <b>Удалить всё расписание целиком</b> (и тогда на сайте про Вашу организацию не будет ничего видно).\n"
    "   - <b>Рекомендуется перед удалением скачать себе всю информацию</b> по расписанию на компьютер, в целях редактирования и отправления нового отредактированного расписания.\n\n"
    "2. <b>Пройдя по кнопке \"Брони/записи\"</b>, Вы сможете:\n"
    "   - <b>Увидеть, кто к Вам уже записался</b> (кто забронировал слот, стол и пр.).\n"
    "   - <b>Отменить запись</b> (если клиент не сможет явиться, например).\n"
    "   - <b>Удалить этот слот</b> (окошко) вообще или <b>добавить новый слот</b> (окно) на то же самое время и дату.\n\n"
)
    
    await callback.message.answer(text, disable_web_page_preview=True, parse_mode="HTML")
    await callback.answer()


@router.callback_query(F.data == "MyBooking")
async def my_offers_main(callback: CallbackQuery):
    """
    Обработка кнопкe "Mои брони/записи".
    .
    """
    tg_id = callback.from_user.id

    #Ищем по тг айди в User этого пользователя и если у него нет numberphone, то показываем кнопку: Привязать номер

    user=await get_user_by_tg_id(tg_id)
    
    # Инициализируем переменную count_booking
    count_booking = 0  # Значение по умолчанию

    # Формируем текст сообщения
    text_massages = (
        f"Здесь можно посмотреть свои записи/брони, которые были сделаны через платформу по ссылке: {url_recording}.\n\n"
    )

    if user.numberphone:
        text_massages += f"У Вас привязан номер телефона: <b>{user.numberphone}</b>\n\n"

        # Получение всех записей по этому номеру:
        filters={
            "numberPhone": user.numberphone
        }
        records=await get_all_records(model_name="Record", filters=filters)
        count_booking=len(records)
        text_massages += f"По этому номер телефона найдено <b>{count_booking} </b> записей/броней.\n\n"
    else:
        text_massages += (
            "У Вас не привязан номер телефона.\n"
            "Чтобы посмотреть свои записи, нужно привязать номер телефона, с которого были сделаны записи/брони."
        )

    # Формируем клавиатуру
    keyboard = kb.create_bookings_my(count_booking=count_booking)

    await callback.message.answer(text_massages, reply_markup=keyboard, disable_web_page_preview=True,parse_mode="HTML")

    await callback.answer()

@router.callback_query(F.data == "numberphoneAdd")
async def bind_phone_handler(callback: CallbackQuery, state: FSMContext):
    """
    Обработка кнопки "Привязать номер телефона".
    """
    await callback.message.answer("Введите номер телефона в формате +79040000000:")
    await state.set_state("waiting_for_phone_booking")  # Устанавливаем состояние ожидания ввода номера
    await callback.answer()

@router.message(StateFilter("waiting_for_phone_booking"))
async def process_phone_input(message: Message, state: FSMContext):
    """
    Обработка ввода номера телефона.
    """
    phone_number = message.text.strip()

    # Проверяем формат номера телефона
    if not phone_number.startswith("+") or not phone_number[1:].isdigit() or len(phone_number) != 12:
        await message.answer("Номер телефона должен быть в формате +79040000000. Пожалуйста, попробуйте еще раз.")
        return

    # Проверяем, чтобы такого номер телефона в базе еще не было
    check=await get_numberPhone(numberPhone=phone_number)
    if check:
        await message.answer(f"Номер телефона {phone_number} уже есть в базе. Свяжитесь с админом, если Вы уверены в правильности предоставления номера телефона!")
    else:

        # Сохраняем номер телефона в базу данных
        tg_id = message.from_user.id
        success = await update_numberphone(tg_id, phone_number)

        if success:
            await message.answer(f"Номер телефона {phone_number} успешно привязан!")
        else:
            await message.answer("Произошла ошибка при привязке номера телефона. Пожалуйста, попробуйте позже.")

    await state.clear()  # Сбрасываем состояние

@router.callback_query(F.data == "viewsBooking")
async def bind_phone_handler(callback: CallbackQuery, state: FSMContext):
    """
    Обработка кнопки "Брони/записи ({count_booking})".
    """
    tg_id = callback.from_user.id

    #Ищем по тг айди в User этого пользователя и если у него нет numberphone, то показываем кнопку: Привязать номер

    user=await get_user_by_tg_id(tg_id)

    if user.numberphone:
        
        # Получение всех записей по этому номеру:
        filters={
            "numberPhone": user.numberphone
        }
        records=await get_all_records(model_name="Record", filters=filters)

        keyboard=kb.info_booking(records)
        await callback.message.answer("Ваши pзаписи/брони:",reply_markup=keyboard)
        
    else:
        text_massages=f"Вам нужно привязать номер телефона"
        await callback.message.answer(text_massages, parse_mode="HTML")
    
    
    await callback.answer()


@router.callback_query(F.data.startswith("booking_"))
async def show_booking_info(callback: CallbackQuery):
    """
    Обработка нажатия на кнопку с информацией о бронировании.
    """
    # Извлекаем id записи из callback_data
    record_id = int(callback.data.split("_")[1])

    
    record=await get_record_by_id(model_name="Record", record_id=record_id)

    if record:
        # Формируем текст сообщения с информацией о бронировании
        message_text = get_message_bookig(record)
        message_text +=f"🔗 Ссылка для брони: {record.url_record_my}\n"
    
        date_booking=record.date_booking

    # Проверяем, является ли запись на сегодня
    today = datetime.now().date()
    if date_booking == today:
        keyboard = kb.yes_or_cancel_booking(record_id)
    else:
        keyboard = kb.cancel_booking(record_id)

    # Отправляем сообщение с информацией о бронировании
    await callback.message.answer(message_text, reply_markup=keyboard, disable_web_page_preview=True,parse_mode="HTML")
    await callback.answer()

@router.callback_query(F.data.startswith("yesBooking_"))
async def yesBooking__booking(callback: CallbackQuery):
    """
    Обработка нажатия на кнопку с информацией о бронировании.
    """
    # Извлекаем id записи из callback_data
    record_id = int(callback.data.split("_")[1])

    
    record=await get_record_by_id(model_name="Record", record_id=record_id)

    if record:
        await update_book_agree_status(record_id, agree_status=True)
        # Формируем текст сообщения с информацией о бронировании
        message_text = "✅ Запись подтверждена пользователем!\n\n"
        message_text += get_message_bookig(record)
        # Отправляем уведомление пользователю с tgId
        try:
            await callback.bot.send_message(
                chat_id=record.tgId,
                text=message_text
            )
        
        except Exception as e:
            logger.error(f"Ошибка при отправке уведомления: {e}")

        
        
        
    # Отправляем сообщение с информацией о бронировании пользователю, который нажал на кнопку
    await callback.message.answer(message_text, parse_mode="HTML")
    await callback.answer()


# Хэндлер для неизвестных команд
@router.message()
async def unknown_command(message: Message):
    await message.answer('Неизвестная команда. Пожалуйста, используйте доступные команды.')